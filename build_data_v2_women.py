"""
build_data_v2_women.py - 女装版重构 (按店铺聚合)
"""
import json, os, math
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook

SRC_DIR = r"D:\周汇报文件"
WOMEN_SHOPS = ["Z-NZTF1店", "G-NZTF1店"]
SHOP_TO_OWNER = {"Z-NZTF1店": "毛立新", "G-NZTF1店": "陈欣诺"}

def load_workbook_safe(path):
    try: return load_workbook(path, data_only=True)
    except Exception as e: print(f"[WARN] {e}"); return None

def parse_date_str(s):
    parts = s.split('.')
    if len(parts)!=3: return None
    y,m,d = parts
    if len(y)==2: y='20'+y
    try: return datetime(int(y),int(m),int(d))
    except: return None

def generate_weeks():
    path = os.path.join(SRC_DIR,"LX利润表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None: raise FileNotFoundError(path)
    ws = wb["分周SKU"]
    ranges, eds = set(), {}
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        dr, ed = row[0], row[1]
        if dr and isinstance(dr,str) and '-' in dr:
            ranges.add(dr)
            if ed and isinstance(ed,datetime): eds[dr]=ed
    sorted_ranges = sorted(ranges, key=lambda x: eds.get(x,datetime(2000,1,1)))
    weeks,iso,labels = [],[],[]
    wy,wm,wym = [],[],[]
    for i,dr in enumerate(sorted_ranges):
        wk=f"W{i+1}"; weeks.append(wk)
        yr,mn = 0,0
        parts=dr.split('-')
        if len(parts)==2:
            edt=parse_date_str(parts[1])
            if edt:
                yr,mn=edt.year,edt.month
                iso.append(f"{edt.year}-W{edt.isocalendar()[1]:02d}")
                labels.append(f"{yr}年{mn:02d}月 W{i+1}·{dr}")
            else: iso.append(f"WK{i+1}"); labels.append(f"{dr} (W{i+1})")
        else: iso.append(f"WK{i+1}"); labels.append(f"{dr} (W{i+1})")
        wy.append(yr); wm.append(mn)
        wym.append(f"{yr}-{mn:02d}" if yr else "")
    print(f"  读取 {len(weeks)} 周: {sorted_ranges[0]} → {sorted_ranges[-1]}")
    return weeks,iso,labels,sorted_ranges,wy,wm,wym

def read_product_list():
    path = os.path.join(SRC_DIR,"产品列表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None: return {},{},{},{}
    ws = wb[wb.sheetnames[0]]
    si,so,sf,sw = {},{},{},{}
    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row[:15]]
        wb_id,sku,img,cat,shop = vals[0],vals[1],vals[4],vals[6],vals[11]
        owner = vals[9] or ""
        create_time = vals[14]  # 创建时间 (datetime)
        if not sku: continue
        sku=str(sku).strip()
        sw[sku]=str(wb_id) if wb_id else ""
        si[sku]=str(img) if img else ""
        so[sku]=str(owner) if owner else ""
        if create_time: sf[sku]=str(create_time)[:10]
    print(f"  产品列表: {len(si)} SKU")
    return si,so,sf,sw

def read_targets():
    path = os.path.join(SRC_DIR,"2026WB年规进度 - 女装.xlsx")
    wb = load_workbook_safe(path)
    if wb is None: return {}
    targets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        mcols = {}
        for c in range(1,ws.max_column+1):
            v = ws.cell(row=2,column=c).value
            if v and isinstance(v,str):
                for mn in ["2月","3月","4月","5月","6月","7月","8月","9月","10月","11月","12月","1月"]:
                    if mn in v: mcols[c]=int(mn.replace("月","")); break
        if not mcols: continue
        trows = {"profit_target":6,"sales_target":10,"real_sales_target":14,"gmv_target":18,"gsv_target":22}
        pd = {}
        for key,rn in trows.items():
            for c,month in mcols.items():
                v = ws.cell(row=rn,column=c).value
                try: v=float(v) if v is not None else 0
                except: v=0
                mk=str(month)
                if mk not in pd: pd[mk]={}
                pd[mk][key]=v
        for mk in pd:
            pd[mk]["margin_target"]=15.0
            pd[mk]["return_target"]=65.0
        targets[sn]=pd
    print(f"  年规目标: {len(targets)} 店铺")
    return targets

def read_profit_data(week_ranges):
    path = os.path.join(SRC_DIR,"LX利润表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None: return {},{},{}
    ws = wb["分周SKU"]
    sw = defaultdict(lambda: defaultdict(lambda: {"gsv":0,"profit":0,"qty":0,"real_sales":0,"products":0}))
    wd = defaultdict(lambda: {"shops":{},"allProducts":[]})
    sa = defaultdict(lambda: defaultdict(lambda: {"gsv":0,"profit":0,"qty":0,"real_sales":0}))
    total = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row)<21: continue
        dr = row[0]
        if not dr or not isinstance(dr,str): continue
        shop = str(row[2] or "")
        if shop not in WOMEN_SHOPS: continue
        sku = str(row[6] or "")
        if not sku: continue
        pf = float(row[8] or 0); gs = float(row[10] or 0)
        rs = float(row[16] or 0); qt = int(row[19] or 0)
        cat = str(row[4] or "")
        mr = float(row[9] or 0)  # 毛利率CNY (decimal)
        rr = float(row[21] or 0) if len(row) > 21 else 0  # 送达退货率 (decimal)
        ad = float(row[35] or 0) if len(row) > 35 else 0  # 广告花费CNY
        # Convert decimal to percentage (same as men's build_data.py)
        margin_pct = round(mr * 100, 2) if mr < 1 else round(mr, 2)
        return_pct = round(rr * 100, 2) if rr < 1 else round(rr, 2)
        total += 1
        sw[shop][dr]["gsv"] += gs; sw[shop][dr]["profit"] += pf
        sw[shop][dr]["real_sales"] += rs; sw[shop][dr]["qty"] += qt
        sw[shop][dr]["products"] += 1
        if shop not in wd[dr]["shops"]:
            wd[dr]["shops"][shop] = {"gsv":0,"profit":0,"margin":0,"products":0,"ad_spend":0}
        sd = wd[dr]["shops"][shop]
        sd["gsv"] += gs; sd["profit"] += pf; sd["products"] += 1
        sd["ad_spend"] += ad
        wd[dr]["allProducts"].append({"sku":sku,"shop":shop,"cat":cat,"profit":round(pf,2),"margin":margin_pct,"gsv":round(gs,2),"qty":qt,"return_rate":return_pct,"ad_spend":round(ad,2),"real_sales":round(rs,2)})
        sa[shop][dr]["gsv"] += gs; sa[shop][dr]["profit"] += pf
        sa[shop][dr]["real_sales"] += rs; sa[shop][dr]["qty"] += qt
    for _,weeks in sw.items():
        for _,d in weeks.items():
            d["margin"] = round(d["profit"]/d["gsv"]*100,2) if d["gsv"]>0 else 0
    for w in wd.values():
        for s in w["shops"].values():
            s["margin"] = round(s["profit"]/s["gsv"]*100,2) if s["gsv"]>0 else 0
        w["top10Profit"] = sorted(w["allProducts"], key=lambda x:-x["profit"])[:10]
    print(f"  利润表: {len(sw)} shops, {len(wd)} weeks, {total} rows")
    return dict(sw), dict(wd), dict(sa)

def read_gmv(week_ranges):
    path = os.path.join(SRC_DIR,"运营日数据.xlsx")
    wb = load_workbook_safe(path)
    if wb is None: return {},{},{}
    ws = wb[wb.sheetnames[0]]
    gd = defaultdict(lambda: defaultdict(float))
    sku_inv = {}  # SKU|shop → inventory
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row)<16: continue
        dv = row[0]; shop = str(row[2] or "")
        if shop not in WOMEN_SHOPS: continue
        gmv = float(row[14] or 0)
        # Extract inventory: col 5 = 可售数量
        sku = str(row[4] or "").strip()
        inv_val = row[5]
        if sku and inv_val is not None:
            try: sku_inv[sku+'|'+shop] = int(inv_val)
            except: pass
        if not dv: continue
        matched = None
        if isinstance(dv,datetime):
            for dr in week_ranges:
                parts=dr.split('-')
                if len(parts)!=2: continue
                sd=parse_date_str(parts[0]); ed=parse_date_str(parts[1])
                if sd and ed and sd<=dv<=ed: matched=dr; break
        if matched: gd[matched][shop] += gmv; total += 1
    # Shop GMV is already shop-level, use directly
    gs = defaultdict(lambda: defaultdict(float))
    for dr,shops in gd.items():
        for shop,gmv in shops.items():
            gs[dr][shop] += gmv
    tw = {}
    for dr,shops in gd.items():
        tw[dr] = {}
        for shop,gmv in shops.items():
            tw[dr][shop] = {"gmv":round(gmv,2),"visitors":0,"atc":0,"qty":0,"click_rate":0,"cart_rate":0,"conv_rate":0}
    print(f"  GMV: {total} entries, inventory: {len(sku_inv)} SKU")
    return tw, dict(gs), sku_inv

def compute_monthly(shop_agg, gmv_shop, targets, week_ranges, wm):
    monthly = defaultdict(lambda: defaultdict(dict))
    # Aggregate actuals by month
    ma = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for dr in week_ranges:
        m = None
        parts=dr.split('-')
        if len(parts)==2:
            dt=parse_date_str(parts[1])
            if dt: m=str(dt.month)
        if not m: continue
        for shop,data in shop_agg.items():
            if dr in data:
                for metric in ["gsv","profit","real_sales","qty"]:
                    ma[shop][m][metric] += data[dr][metric]
        if dr in gmv_shop:
            for shop,gmv in gmv_shop[dr].items():
                ma[shop][m]["gmv"] += gmv

    today6 = datetime(datetime.now().year,datetime.now().month,datetime.now().day)
    for shop in targets:
        monthly[shop] = {}
        for month in range(2,13):
            mk = str(month)
            year = 2027 if month<=1 else 2026
            ms = datetime(year,month,1)
            me = datetime(year,month+1,1)-timedelta(days=1) if month<12 else datetime(year,12,31)
            if today6>me: tp=100.0
            elif today6<ms: tp=0.0
            else: tp=round((today6-ms).days/(me-ms).days*100,1)

            a = ma.get(shop,{}).get(mk,{})
            t = targets.get(shop,{}).get(mk,{})
            ap,ags,ar,ag,aq = a.get("profit",0),a.get("gsv",0),a.get("real_sales",0),a.get("gmv",0),a.get("qty",0)

            entry = {
                "time_progress":tp,
                "actual":{"profit":round(ap,2),"gsv":round(ags,2),"real_sales":round(ar,2),"qty":int(aq),"gmv":round(ag,2)},
                "target":{"profit_target":round(t.get("profit_target",0),2),"sales_target":int(t.get("sales_target",0)),
                          "real_sales_target":round(t.get("real_sales_target",0),2),"gmv_target":round(t.get("gmv_target",0),2),
                          "gsv_target":round(t.get("gsv_target",0),2),"margin_target":15.0,"return_target":65.0},
                "derived":{"margin_pct":round(ap/ags*100,2) if ags>0 else "-","return_rate":round((1-(ags*12.5)/ag)*100,2) if ag>0 else "-"},
            }
            def prog(av,tv):
                if tv<=0 or tp<=0: return "-"
                return round((av/tv)*100/tp*100,1)
            entry["completion"] = {"profit_progress":prog(ap,t.get("profit_target",0)),"sales_progress":prog(aq,t.get("sales_target",0)),
                                    "real_sales_progress":prog(ar,t.get("real_sales_target",0)),"gmv_progress":prog(ag,t.get("gmv_target",0)),
                                    "gsv_progress":prog(ags,t.get("gsv_target",0))}
            monthly[shop][mk] = entry
    return dict(monthly)

def main():
    print("="*60); print("build_data_v2_women.py - 女装版重构"); print("="*60)
    print("\n[1/5] 周数组...")
    weeks,iso,labels,ranges,wy,wm,wym = generate_weeks()
    print("\n[2/5] 产品列表...")
    si,so,sf,sw = read_product_list()
    print("\n[3/5] 年规目标...")
    targets = read_targets()
    print("\n[4/5] 利润表 + GMV...")
    shop_weekly, week_data, shop_agg = read_profit_data(ranges)
    traffic_weekly, gmv_shop, sku_inventory = read_gmv(ranges)
    print("\n[5/5] 完成进度...")
    monthly = compute_monthly(shop_agg, gmv_shop, targets, ranges, wm)

    # Build SHOP_OWNERS and owner-keyed targets
    all_shops = set(list(targets.keys()) + list(shop_weekly.keys()) + list(WOMEN_SHOPS))
    shop_owners_out = {}
    for shop in all_shops:
        owner = SHOP_TO_OWNER.get(shop, "其他/待定")
        shop_owners_out[shop] = {owner: True}

    # Restructure PERSON_TARGETS to be owner-keyed
    owner_targets = {}
    for shop, st in targets.items():
        owner = SHOP_TO_OWNER.get(shop, "其他/待定")
        if owner not in owner_targets:
            owner_targets[owner] = {}
        for month, mdata in st.items():
            if month not in owner_targets[owner]:
                owner_targets[owner][month] = {"profit_target":0,"sales_target":0,"real_sales_target":0,"gmv_target":0,"gsv_target":0,"margin_target":15.0,"return_target":65.0}
            for key in ["profit_target","sales_target","real_sales_target","gmv_target","gsv_target"]:
                owner_targets[owner][month][key] += mdata.get(key, 0)

    # Build OWNERS list (owners with targets first, then others)
    owners_list = sorted([o for o in owner_targets if o != "其他/待定"])
    if "其他/待定" in owner_targets:
        owners_list.append("其他/待定")
    # Also add owners from SHOP_TO_OWNER that might not have targets yet
    for o in SHOP_TO_OWNER.values():
        if o not in owners_list and o not in owner_targets:
            owners_list.append(o)

    # Build SKU_OWNER_LOOKUP (already in correct format from read_product_list)
    sku_owner_lookup = so

    ki = [i for i,yr in enumerate(wy) if yr>=2025]
    w2 = [weeks[i] for i in ki]; wi2 = [iso[i] for i in ki]; wl2 = [labels[i] for i in ki]
    wr2 = [ranges[i] for i in ki]; wy2 = [wy[i] for i in ki]; wm2 = [wm[i] for i in ki]; wym2 = [wym[i] for i in ki]
    sw2 = {s:{k:v for k,v in w.items() if k in wr2} for s,w in shop_weekly.items()}
    wd2 = {k:v for k,v in week_data.items() if k in wr2}
    tw2 = {k:v for k,v in traffic_weekly.items() if k in wr2}

    data = {"WEEKS":w2,"WEEKS_ISO":wi2,"WEEK_LABELS":wl2,"WEEK_RANGES":wr2,
            "WEEK_YEARS":wy2,"WEEK_MONTHS":wm2,"WEEK_YEAR_MONTHS":wym2,
            "SHOP_WEEKLY":sw2,"WEEK_DATA":wd2,"TRAFFIC_WEEKLY":tw2,
            "PERSON_TARGETS":owner_targets,"MONTHLY_COMPLETION":monthly,
            "SKU_IMG":si,"SKU_OWNER":so,"SKU_FIRST_DATE":sf,"SKU_WB_ID":sw,
            "SKU_INVENTORY":sku_inventory,
            "SHOP_OWNERS":shop_owners_out,"OWNERS":owners_list,"SKU_OWNER_LOOKUP":so,
            "_VERSION":"v2.0-women"}

    def sanitize(obj):
        if isinstance(obj,float): return None if (math.isnan(obj) or math.isinf(obj)) else obj
        if isinstance(obj,dict): return {k:sanitize(v) for k,v in obj.items()}
        if isinstance(obj,list): return [sanitize(v) for v in obj]
        return obj

    dc = sanitize(data)
    js = json.dumps(dc, ensure_ascii=False, default=str)
    out = os.path.join(SRC_DIR,"data-detail-women.js")
    with open(out,"w",encoding="utf-8") as f: f.write(f"var DATA = {js};")
    print(f"\n输出: data-detail-women.js ({os.path.getsize(out)/1024/1024:.1f} MB)")
    print(f"  周:{len(w2)} 店铺:{len(targets)} 负责人:{len(owners_list)} 数据店铺:{len(sw2)}")
    print("完成!")

if __name__=="__main__": main()
