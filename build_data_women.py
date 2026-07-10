#!/usr/bin/env python3
"""
build_data_women.py - 女装版：从源Excel文件生成 data_women.js
源文件:
  - D:/周汇报文件/运营日数据.xlsx
  - D:/周汇报文件/产品列表.xlsx
  - D:/周汇报文件/LX利润表.xlsx
  - D:/周汇报文件/2026WB年规进度 - 女装.xlsx
输出: data_women.js
"""

import json
import math
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl

# ── 配置 ──────────────────────────────────────────────
SRC_DIR = r"D:\周汇报文件"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "data_women.js")

# ── 工具函数 ──────────────────────────────────────────

def sanitize_value(v):
    """确保值是合法JSON：NaN/Inf/-Inf → null，字符串数字 → float"""
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        return round(v, 6) if abs(v) < 10 else round(v, 2)
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, str):
        v_stripped = v.strip()
        if not v_stripped:
            return None
        try:
            f = float(v_stripped)
            if math.isnan(f) or math.isinf(f):
                return None
            return round(f, 6) if abs(f) < 10 else round(f, 2)
        except (ValueError, TypeError):
            return v
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return str(v)


def iso_week_to_date_range(iso_week_str):
    """'2025-W30' → (monday, sunday)"""
    year_s, week_s = iso_week_str.split("-W")
    year, week = int(year_s), int(week_s)
    jan4 = date(year, 1, 4)
    monday = jan4 - timedelta(days=jan4.weekday()) + timedelta(weeks=week - 1)
    return monday, monday + timedelta(days=6)


def date_to_iso_week(d):
    """date → '2025-W30'"""
    if isinstance(d, datetime):
        d = d.date()
    iso = d.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def load_workbook_safe(path):
    """安全加载workbook"""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return wb
    except Exception as e:
        print(f"ERROR: 无法加载 {path}: {e}")
        return None


# ── 阶段 1: 生成WEEKS等基础数据 ───────────────────────

def generate_weeks():
    """从 2026.01.26(周一) 动态生成至今天所在周的数组"""
    from datetime import date as dt_date, timedelta
    start_date = dt_date(2026, 1, 26)  # W1 = 2026.01.26-2026.02.01
    today = dt_date.today()
    # 计算从 start_date 到今天经过了多少周
    delta_days = (today - start_date).days
    if delta_days < 0:
        delta_days = 0
    total_weeks = (delta_days // 7) + 1

    weeks = []
    weeks_iso = []
    week_labels = []
    for i in range(total_weeks):
        week_start = start_date + timedelta(weeks=i)
        week_end = week_start + timedelta(days=6)
        # ISO week of the Monday
        iso_year, iso_wn, _ = week_start.isocalendar()
        iso = f"{iso_year}-W{iso_wn:02d}"
        label = f"W{i+1} ({week_start.month:02d}.{week_start.day:02d}-{week_end.month:02d}.{week_end.day:02d})"
        weeks.append(f"W{i+1}")
        weeks_iso.append(iso)
        week_labels.append(label)
    return weeks, weeks_iso, week_labels


# ── 阶段 2: 读取产品列表 ──────────────────────────────

def read_product_list():
    """从产品列表.xlsx 提取 SKU映射"""
    path = os.path.join(SRC_DIR, "产品列表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None:
        return {}, {}, {}, {}, {}

    ws = wb[wb.sheetnames[0]]
    sku_img = {}
    sku_owner = {}
    sku_first_date = {}
    sku_wb_id = {}
    shop_owner_set = defaultdict(set)

    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if i > 2000:
            break
        vals = [cell.value for cell in row[:15]]
        wb_id = vals[0]
        sku = vals[1]
        img = vals[9]
        shop = vals[10] if len(vals) > 10 else None
        owner = vals[13] if len(vals) > 13 else None
        create_time = vals[14] if len(vals) > 14 else None

        if not sku or not wb_id:
            continue

        wb_id_str = str(int(wb_id)) if isinstance(wb_id, float) else str(wb_id)
        
        sku_wb_key = f"{sku}|{wb_id_str}"
        if shop:
            sku_wb_key_shop = f"{sku}|{shop}|{wb_id_str}"
        
        if img:
            sku_img[sku_wb_key] = str(img)
            if shop:
                sku_img[sku_wb_key_shop] = str(img)
        if owner:
            sku_owner[sku_wb_key] = str(owner)
            if shop:
                shop_owner_set[str(shop)].add(str(owner))
                sku_owner[sku_wb_key_shop] = str(owner)
        if create_time:
            fd_str = None
            if isinstance(create_time, (datetime, date)):
                fd_str = create_time.strftime("%Y-%m-%d") if isinstance(create_time, datetime) else create_time.isoformat()
            elif isinstance(create_time, str):
                fd_str = create_time[:10]
            if fd_str:
                sku_first_date[sku_wb_key] = fd_str
                if shop:
                    sku_first_date[sku_wb_key_shop] = fd_str
        
        sku_wb_id[sku_wb_key] = wb_id_str
        if shop:
            sku_wb_id[sku_wb_key_shop] = wb_id_str

    shop_owners = {s: {o: True for o in owners} for s, owners in shop_owner_set.items()}

    wb.close()
    print(f"  产品列表: {len(sku_img)} SKU图片, {len(sku_owner)} SKU负责人, "
          f"{len(shop_owners)} 店铺负责人, {len(sku_wb_id)} SKU-WB映射, {len(sku_first_date)} 首次日期")
    return sku_img, sku_owner, sku_first_date, sku_wb_id, shop_owners


# ── 阶段 3: 读取LX利润表 ──────────────────────────────

def read_lx_profit(weeks_iso):
    """从LX利润表.xlsx 生成 WEEK_DATA 和 SHOP_WEEKLY"""
    path = os.path.join(SRC_DIR, "LX利润表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None:
        return {}, {}

    iso_to_w = {iso: f"W{i+1}" for i, iso in enumerate(weeks_iso)}

    ws_shop = wb["店铺分周利润表"]
    shop_weekly = defaultdict(dict)

    for i, row in enumerate(ws_shop.iter_rows(min_row=3)):
        if i > 5000:
            break
        vals = [cell.value for cell in row[:10]]
        week_end = vals[1]
        shop_name = vals[2]
        margin_val = vals[4]
        gsv_val = vals[6]

        if not week_end or not shop_name:
            continue

        iso = date_to_iso_week(week_end) if isinstance(week_end, (datetime, date)) else str(week_end)
        w_key = iso_to_w.get(iso)
        if w_key is None:
            continue

        margin = sanitize_value(margin_val)
        gsv = sanitize_value(gsv_val)

        if margin is not None and gsv is not None:
            shop_weekly[str(shop_name)][w_key] = {
                "gsv": gsv,
                "margin": round(margin * 100, 4) if isinstance(margin, float) and margin < 1 else margin
            }

    shop_weekly = {k: dict(v) for k, v in shop_weekly.items()}

    ws_sku = wb["分周SKU"]
    week_data_raw = defaultdict(list)

    sku_count = 0
    for i, row in enumerate(ws_sku.iter_rows(min_row=3)):
        if i > 40000:
            break
        vals = [cell.value for cell in row[:37]]
        week_end = vals[1]
        shop = vals[2]
        cat = vals[4]
        sku = vals[6]
        profit = vals[8]
        margin_rate = vals[9]
        gsv = vals[10]
        qty = vals[14]
        return_rate = vals[21] if len(vals) > 21 else None
        ad_spend = vals[35] if len(vals) > 35 else None
        unit_delivery = vals[17] if len(vals) > 17 else None  # 单件尾程配送CNY
        unit_return_fee = vals[18] if len(vals) > 18 else None  # 单件退货费CNY

        if not week_end or not sku:
            continue

        iso = date_to_iso_week(week_end) if isinstance(week_end, (datetime, date)) else str(week_end)
        w_key = iso_to_w.get(iso)
        if w_key is None:
            continue

        sku_str = str(sku).strip() if sku else ""

        product = {
            "sku": sku_str or "",
            "shop": str(shop) if shop else "",
            "cat": str(cat) if cat else "",
            "profit": sanitize_value(profit) or 0,
            "margin": sanitize_value(margin_rate),
            "gsv": sanitize_value(gsv) or 0,
            "qty": sanitize_value(qty) or 0,
            "return_rate": sanitize_value(return_rate),
            "ad_spend": sanitize_value(ad_spend) or 0,
            "unit_delivery": sanitize_value(unit_delivery) or 0,
            "unit_return_fee": sanitize_value(unit_return_fee) or 0
        }

        if isinstance(product["margin"], float) and product["margin"] < 1:
            product["margin"] = round(product["margin"] * 100, 2)
        if isinstance(product["return_rate"], float) and product["return_rate"] < 1:
            product["return_rate"] = round(product["return_rate"] * 100, 2)

        week_data_raw[w_key].append(product)
        sku_count += 1

    week_data = {}
    for w_key in sorted(week_data_raw.keys(), key=lambda x: int(x[1:])):
        products = week_data_raw[w_key]
        products.sort(key=lambda p: p["profit"], reverse=True)

        shop_summary = defaultdict(lambda: {"gsv": 0, "profit": 0, "margin": 0, "products": 0, "ad_spend": 0})
        for p in products:
            s = shop_summary[p["shop"]]
            s["gsv"] += p["gsv"]
            s["profit"] += p["profit"]
            s["products"] += 1
            s["ad_spend"] += p.get("ad_spend", 0) or 0

        for s in shop_summary.values():
            if s["gsv"] > 0:
                s["margin"] = round(s["profit"] / s["gsv"] * 100, 2)
            else:
                s["margin"] = 0

        top10 = [p for p in products if p["sku"] != "无匹配ID费用"][:10]

        week_data[w_key] = {
            "shops": {k: dict(v) for k, v in shop_summary.items()},
            "top10Profit": top10,
            "allProducts": products
        }

    wb.close()
    print(f"  SHOP_WEEKLY: {len(shop_weekly)} shops")
    print(f"  WEEK_DATA: {len(week_data)} weeks, {sku_count} total product-weeks")
    return shop_weekly, week_data


# ── 阶段 4: 读取运营日数据 → TRAFFIC_WEEKLY ──────────

def read_traffic_weekly(weeks_iso):
    """从运营日数据.xlsx 生成 TRAFFIC_WEEKLY"""
    path = os.path.join(SRC_DIR, "运营日数据.xlsx")
    wb = load_workbook_safe(path)
    if wb is None:
        return {}

    iso_to_w = {iso: f"W{i+1}" for i, iso in enumerate(weeks_iso)}
    ws = wb[wb.sheetnames[0]]

    weekly_agg = defaultdict(lambda: defaultdict(lambda: {
        "visitors": 0, "atc": 0, "qty": 0,
        "click_cnt": 0, "click_impressions": 0,
        "return_qty": 0, "total_qty_ref": 0
    }))

    sku_first_inventory_date = {}
    sku_latest_inventory = {}

    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if i > 200000:
            break
        if not row:
            continue

        d_val = row[0].value if len(row) > 0 else None
        sku = row[4].value if len(row) > 4 else None          # 卖家SKU (col 4)
        inventory = row[5].value if len(row) > 5 else None     # 可售数量 (col 5)
        visitors = row[8].value if len(row) > 8 else None      # 访客 (col 8)
        atc = row[11].value if len(row) > 11 else None         # 加购数 (col 11)
        qty = row[13].value if len(row) > 13 else None         # 销量 (col 13)
        click_rate_val = row[27].value if len(row) > 27 else None  # 广告点击率 (col 27)
        return_rate_raw = row[23].value if len(row) > 23 else None # 财报退货率 (col 23)

        if not d_val or not sku:
            continue

        if isinstance(d_val, datetime):
            d_date = d_val.date()
        elif isinstance(d_val, date):
            d_date = d_val
        elif isinstance(d_val, str):
            try:
                d_date = datetime.strptime(d_val[:10], "%Y-%m-%d").date()
            except:
                continue
        else:
            continue

        iso = date_to_iso_week(d_date)
        w_key = iso_to_w.get(iso)
        if w_key is None:
            continue

        sku_str = str(sku).strip()
        agg = weekly_agg[w_key][sku_str]
        agg["visitors"] += float(visitors) if visitors else 0
        agg["atc"] += float(atc) if atc else 0
        agg["qty"] += float(qty) if qty else 0

        if inventory is not None:
            try:
                inv_val = float(inventory)
                if inv_val > 0:
                    if sku_str not in sku_first_inventory_date or d_date < sku_first_inventory_date[sku_str]:
                        sku_first_inventory_date[sku_str] = d_date
                if sku_str not in sku_latest_inventory or d_date > sku_latest_inventory[sku_str]['date']:
                    sku_latest_inventory[sku_str] = {'date': d_date, 'value': inv_val}
            except (ValueError, TypeError):
                pass

        if click_rate_val is not None and visitors:
            click_rate_f = float(click_rate_val)
            if not (math.isnan(click_rate_f) or math.isinf(click_rate_f)):
                agg["click_cnt"] += click_rate_f * float(visitors)
                agg["click_impressions"] += float(visitors)

    traffic_weekly = {}
    for w_key, sku_data in sorted(weekly_agg.items(), key=lambda x: int(x[0][1:])):
        traffic_weekly[w_key] = {}
        for sku, agg in sku_data.items():
            v = agg["visitors"]
            if agg["click_impressions"] > 0:
                click_rate = round(agg["click_cnt"] / agg["click_impressions"], 6)
            else:
                click_rate = None
            atc_rate = round(agg["atc"] / v, 6) if v > 0 else 0.0
            conv_rate = round(agg["qty"] / v, 6) if v > 0 else 0.0
            return_rate = None
            sales_qty = round(agg["qty"], 0)
            traffic_weekly[w_key][sku] = [click_rate, atc_rate, conv_rate, return_rate, sales_qty]

    wb.close()
    total_entries = sum(len(v) for v in traffic_weekly.values())
    print(f"  TRAFFIC_WEEKLY: {len(traffic_weekly)} weeks, {total_entries} total SKU-week entries")
    print(f"  SKU首次库存日期 (运营日数据): {len(sku_first_inventory_date)} 个SKU")
    print(f"  SKU最新可售数量 (运营日数据): {len(sku_latest_inventory)} 个SKU")
    return traffic_weekly, sku_first_inventory_date, sku_latest_inventory


# ── 阶段 5: 读取年规进度 → PERSON_TARGETS ─────────────

# 女装版店铺→负责人映射
WOMEN_SHOP_TO_OWNER = {
    "Z-NZTF1店": "毛立新",
    "G-NZTF1店": "陈欣诺",
    "WB纯白关店": "其他/待定",
    "OZ女装店": "其他/待定",
    "WB汤总女装店": "其他/待定",
}

def read_person_targets():
    """从2026WB年规进度 - 女装.xlsx 生成 PERSON_TARGETS（支持同一负责人多店铺合并）"""
    path = os.path.join(SRC_DIR, "2026WB年规进度 - 女装.xlsx")
    wb = load_workbook_safe(path)
    if wb is None:
        return {}

    col_to_month = {2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7, 8: 8,
                    9: 9, 10: 10, 11: 11, 12: 12, 13: 1}

    all_targets = {}

    def parse_shop_sheet(ws):
        """解析单个店铺 sheet，返回 {month: {gsv_target, gsv_done, ...}}"""
        rows_data = {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100)):
            vals = [cell.value for cell in row[:16]]
            rows_data[i+1] = vals

        targets_by_month = defaultdict(dict)

        shop_section_started = False
        for row_num, vals in rows_data.items():
            a_val = vals[0] if len(vals) > 0 else None
            if a_val and isinstance(a_val, str) and "店" in a_val:
                shop_section_started = True
                continue
            if shop_section_started:
                continue
            label = str(vals[1]).strip() if vals[1] is not None else ""
            if not label:
                continue

            if "月度目标" in label and "利润" in label:
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["profit_target"] = sanitize_value(v) or 0

            elif "实际利润" in label and "率" not in label:
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["profit_done"] = sanitize_value(v) or 0

            elif label == "销量目标":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["sales_target"] = sanitize_value(v) or 0

            elif label == "销量完成":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["sales_done"] = sanitize_value(v) or 0

            elif label == "GMV目标":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["gmv_target"] = sanitize_value(v) or 0

            elif label == "GMV完成":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["gmv_done"] = sanitize_value(v) or 0

            elif label == "GSV目标":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["gsv_target"] = sanitize_value(v) or 0

            elif label == "GSV完成":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["gsv_done"] = sanitize_value(v) or 0

        return targets_by_month

    def finalize_targets(targets_by_month):
        """补全 1-12 月 + 计算利润率"""
        for month in targets_by_month:
            t = targets_by_month[month]
            gsv_done = t.get("gsv_done", 0)
            profit_done = t.get("profit_done", 0)
            if gsv_done and gsv_done > 0 and profit_done:
                t["profit_rate"] = round(profit_done / gsv_done * 100, 2)
            else:
                t["profit_rate"] = 0
        for m in range(1, 13):
            if m not in targets_by_month:
                targets_by_month[m] = {
                    "gsv_target": 0, "gsv_done": 0,
                    "gmv_target": 0, "gmv_done": 0,
                    "sales_target": 0, "sales_done": 0,
                    "profit_target": 0, "profit_done": 0,
                    "profit_rate": 0
                }
        return {str(m): dict(targets_by_month[m]) for m in range(1, 13)}

    def merge_targets(existing, incoming):
        """将 incoming 的月度数据累加到 existing 上"""
        for m_str in existing:
            em = existing[m_str]
            im = incoming.get(m_str, {})
            for k in em:
                em[k] = em[k] + (im.get(k, 0) or 0)
        return existing

    for sheet_name in wb.sheetnames:
        owner_name = WOMEN_SHOP_TO_OWNER.get(sheet_name)
        if owner_name is None:
            continue
        ws = wb[sheet_name]
        shop_targets = parse_shop_sheet(ws)
        person_data = finalize_targets(shop_targets)

        if owner_name in all_targets:
            # 同一负责人多个店铺：累加
            all_targets[owner_name] = merge_targets(all_targets[owner_name], person_data)
        else:
            all_targets[owner_name] = person_data

    wb.close()
    print(f"  PERSON_TARGETS: {len(all_targets)} people")
    return all_targets


# ── 阶段 5.5: 全量预计算新品等级和销售等级 ──────────────

def compute_new_product_grades(week_data, sku_first_date, weeks_iso):
    all_products = set()
    for wk, wd in week_data.items():
        for p in wd.get("allProducts", []):
            shop = p.get("shop", "")
            sku = p.get("sku", "")
            if shop and sku and sku != "无匹配ID费用":
                all_products.add(f"{shop}|{sku}")

    iso_to_dates = {}
    for iso in weeks_iso:
        start, end = iso_week_to_date_range(iso)
        iso_to_dates[iso] = (start, end)

    iso_to_w = {iso: f"W{i+1}" for i, iso in enumerate(weeks_iso)}

    today = date.today()
    new_product_grades = {}

    for key in all_products:
        parts = key.split("|", 1)
        shop = parts[0]
        sku = parts[1] if len(parts) > 1 else ""

        fd = None
        for lk in [key, f"{sku}|{shop}", sku]:
            if lk in sku_first_date:
                fd = sku_first_date[lk]
                break

        if not fd:
            new_product_grades[key] = None
            continue

        try:
            parts_fd = fd.split("-")
            first_date = date(int(parts_fd[0]), int(parts_fd[1]), int(parts_fd[2]))
        except (ValueError, IndexError):
            new_product_grades[key] = None
            continue

        cutoff_date = first_date + timedelta(days=45)

        if first_date > today:
            new_product_grades[key] = None
            continue

        total_qty = 0.0
        for iso in weeks_iso:
            ws, we = iso_to_dates[iso]
            if we >= first_date and ws <= cutoff_date:
                w_key = iso_to_w.get(iso)
                if w_key and w_key in week_data:
                    for p in week_data[w_key].get("allProducts", []):
                        if p.get("shop") == shop and p.get("sku") == sku:
                            total_qty += p.get("qty", 0) or 0

        if total_qty >= 175:
            new_product_grades[key] = "S"
        elif total_qty >= 84:
            new_product_grades[key] = "A"
        elif total_qty >= 56:
            new_product_grades[key] = "B"
        else:
            new_product_grades[key] = None

    s_count = sum(1 for v in new_product_grades.values() if v == "S")
    a_count = sum(1 for v in new_product_grades.values() if v == "A")
    b_count = sum(1 for v in new_product_grades.values() if v == "B")
    print(f"  NEW_PRODUCT_GRADE (按销量): {len(new_product_grades)} SKUs (S={s_count}, A={a_count}, B={b_count})")
    return new_product_grades


def compute_sales_grades(week_data, sku_first_date, weeks_iso, sku_inventory):
    all_products = set()
    for wk, wd in week_data.items():
        for p in wd.get("allProducts", []):
            shop = p.get("shop", "")
            sku = p.get("sku", "")
            if shop and sku and sku != "无匹配ID费用":
                all_products.add(f"{shop}|{sku}")

    sales_grades = {}

    for key in all_products:
        parts = key.split("|", 1)
        shop = parts[0]
        sku = parts[1] if len(parts) > 1 else ""

        max_weekly_qty = 0.0
        for wk, wd in week_data.items():
            for p in wd.get("allProducts", []):
                if p.get("shop") == shop and p.get("sku") == sku:
                    qty = p.get("qty", 0) or 0
                    if qty > max_weekly_qty:
                        max_weekly_qty = qty

        if max_weekly_qty >= 350:
            sales_grades[key] = "S"
        elif max_weekly_qty >= 175:
            sales_grades[key] = "A"
        elif max_weekly_qty >= 105:
            sales_grades[key] = "B"
        else:
            sales_grades[key] = None

    s_count = sum(1 for v in sales_grades.values() if v == "S")
    a_count = sum(1 for v in sales_grades.values() if v == "A")
    b_count = sum(1 for v in sales_grades.values() if v == "B")
    none_count = sum(1 for v in sales_grades.values() if v is None)
    print(f"  SALES_GRADE (单周最高销量): {len(sales_grades)} SKUs (S={s_count}, A={a_count}, B={b_count}, None={none_count})")
    return sales_grades


# ── 主函数 ────────────────────────────────────────────

def main():
    print("=" * 60)
    print("build_data_women.py - 生成 data_women.js (女装版)")
    print("=" * 60)

    # 阶段 1: 基础周数组
    print("\n[1/5] 生成周数组...")
    weeks, weeks_iso, week_labels = generate_weeks()
    print(f"  共 {len(weeks)} 周: {weeks_iso[0]} → {weeks_iso[-1]}")

    # 阶段 2: 产品列表
    print("\n[2/5] 读取产品列表...")
    sku_img, sku_owner, sku_first_date, sku_wb_id, shop_owners = read_product_list()

    # 阶段 3: 利润表
    print("\n[3/5] 读取LX利润表...")
    shop_weekly, week_data = read_lx_profit(weeks_iso)

    # 阶段 4: 运营日数据
    print("\n[4/5] 读取运营日数据...")
    traffic_weekly, sku_first_inventory_date, sku_latest_inventory = read_traffic_weekly(weeks_iso)

    print("  合并运营日数据 M列(销量) 到 WEEK_DATA...")
    qty_merged = 0
    for w_key in week_data:
        if w_key not in traffic_weekly:
            continue
        sku_qty = traffic_weekly[w_key]
        for p in week_data[w_key]["allProducts"]:
            sku = p["sku"]
            if sku in sku_qty and len(sku_qty[sku]) > 4:
                new_qty = sku_qty[sku][4]
                if new_qty and new_qty > 0:
                    p["qty"] = new_qty
                    qty_merged += 1
    print(f"  已合并 {qty_merged} 条销量数据 (来源: 运营日数据 M列)")

    sku_to_wb_ids = {}
    for key in sku_wb_id:
        parts = key.split('|')
        if len(parts) >= 2:
            sku = parts[0]
            wb_id = sku_wb_id[key]
            if sku not in sku_to_wb_ids:
                sku_to_wb_ids[sku] = set()
            sku_to_wb_ids[sku].add(wb_id)
    
    merged_sku_first_date = {}
    
    for sku, d in sku_first_inventory_date.items():
        date_str = d.strftime("%Y-%m-%d") if isinstance(d, date) else str(d)[:10]
        merged_sku_first_date[sku] = date_str
        if sku in sku_to_wb_ids:
            for wb_id in sku_to_wb_ids[sku]:
                merged_sku_first_date[f"{sku}|{wb_id}"] = date_str
    
    for key, date_str in sku_first_date.items():
        if key not in merged_sku_first_date:
            merged_sku_first_date[key] = date_str
    
    shop_sku_added = 0
    for wk, wd in week_data.items():
        for p in wd.get("allProducts", []):
            shop = p.get("shop", "")
            sku = p.get("sku", "")
            if not shop or not sku or sku == "无匹配ID费用":
                continue
            shop_sku_key = f"{shop}|{sku}"
            if shop_sku_key in merged_sku_first_date:
                continue
            fd = None
            for candidate in [sku, f"{sku}|{shop}"]:
                if candidate in merged_sku_first_date:
                    fd = merged_sku_first_date[candidate]
                    break
            if not fd and sku in sku_to_wb_ids:
                for wb_id in sku_to_wb_ids[sku]:
                    for candidate in [f"{sku}|{wb_id}", f"{sku}|{shop}|{wb_id}"]:
                        if candidate in merged_sku_first_date:
                            fd = merged_sku_first_date[candidate]
                            break
                    if fd:
                        break
            if fd:
                merged_sku_first_date[shop_sku_key] = fd
                shop_sku_added += 1

    print(f"  SKU_FIRST_DATE: {len(merged_sku_first_date)} 个 (合并产品列表+运营日数据), 新增 shop|sku 映射 {shop_sku_added} 个")
    print(f"    无库存记录的SKU将不显示上架天数（显示为 '-'）")

    sku_inventory = {}
    for sku, info in sku_latest_inventory.items():
        sku_inventory[sku] = int(info['value'])

    # 添加 shop|sku 映射，让渲染代码能通过 SKU代码|店铺 键查找库存
    inv_shop_sku_added = 0
    for wk, wd in week_data.items():
        for p in wd.get("allProducts", []):
            shop = p.get("shop", "")
            sku = p.get("sku", "")
            if not shop or not sku or sku == "无匹配ID费用":
                continue
            shop_sku_key = f"{shop}|{sku}"
            if shop_sku_key in sku_inventory:
                continue
            if sku in sku_to_wb_ids:
                for wb_id in sku_to_wb_ids[sku]:
                    wb_id_str = str(wb_id)
                    if wb_id_str in sku_inventory:
                        sku_inventory[shop_sku_key] = sku_inventory[wb_id_str]
                        inv_shop_sku_added += 1
                        break

    print(f"  SKU_INVENTORY: {len(sku_inventory)} 个SKU (最新日期可售数量), 新增 shop|sku 映射 {inv_shop_sku_added} 个")

    # 阶段 5: 年规进度 (女装版)
    print("\n[5/5] 读取年规进度 (女装)...")
    person_targets = read_person_targets()

    sku_owner_lookup = {}
    for key, owner in sku_owner.items():
        parts = key.split('|')
        if len(parts) == 3:
            lookup_key = f"{parts[0]}|{parts[1]}"
            sku_owner_lookup[lookup_key] = owner
        if parts[0] not in sku_owner_lookup:
            sku_owner_lookup[parts[0]] = owner
    print(f"  SKU_OWNER_LOOKUP: {len(sku_owner_lookup)} entries")

    # 女装版：G-NZTF1店 全部产品强制归属 → 陈欣诺（产品列表Excel中可能是江凯伦/陈敏华等）
    remap_count = 0
    for key in list(sku_owner_lookup.keys()):
        if 'G-NZTF1店' in key:
            sku_owner_lookup[key] = '陈欣诺'
            remap_count += 1
    # 同步修正 SKU_OWNER 原始数据（HTML 也有从 SKU_OWNER 重建 fallback 的路径）
    for key in list(sku_owner.keys()):
        if 'G-NZTF1店' in key:
            sku_owner[key] = '陈欣诺'
    # 同步修正 SHOP_OWNERS
    if 'G-NZTF1店' in shop_owners:
        shop_owners['G-NZTF1店'] = {'陈欣诺': True}
    # 三个无数据店铺挂到"其他/待定"
    shop_owners['WB纯白关店'] = {'其他/待定': True}
    shop_owners['OZ女装店'] = {'其他/待定': True}
    shop_owners['WB汤总女装店'] = {'其他/待定': True}
    print(f"  G-NZTF1店 负责人强制归属陈欣诺: {remap_count} 条")

    # ── 女装版：过滤到仅女装5店 ──
    women_shops = set(WOMEN_SHOP_TO_OWNER.keys())
    print(f"\n  [女装过滤] 仅保留 {len(women_shops)} 家店铺: {women_shops}")
    
    # 1. SHOP_WEEKLY
    orig_sw = len(shop_weekly)
    shop_weekly = {k: v for k, v in shop_weekly.items() if k in women_shops}
    print(f"    SHOP_WEEKLY: {orig_sw} → {len(shop_weekly)}")
    
    # 2. WEEK_DATA: filter shops and allProducts
    orig_wd_products = sum(len(wd.get('allProducts', [])) for wd in week_data.values())
    for wk in list(week_data.keys()):
        wd = week_data[wk]
        wd['shops'] = {k: v for k, v in wd.get('shops', {}).items() if k in women_shops}
        wd['allProducts'] = [p for p in wd.get('allProducts', []) if p.get('shop') in women_shops]
    new_wd_products = sum(len(wd.get('allProducts', [])) for wd in week_data.values())
    print(f"    WEEK_DATA products: {orig_wd_products} → {new_wd_products}")
    
    # 3. SHOP_OWNERS
    orig_so = len(shop_owners)
    shop_owners = {k: v for k, v in shop_owners.items() if k in women_shops}
    print(f"    SHOP_OWNERS: {orig_so} → {len(shop_owners)}")

    print("\n[5.5] 预计算新品等级和销售等级...")
    new_product_grades = compute_new_product_grades(
        week_data, merged_sku_first_date, weeks_iso
    )
    sales_grades = compute_sales_grades(
        week_data, merged_sku_first_date, weeks_iso, sku_inventory
    )

    print("\n" + "=" * 60)
    print("组装 data_women.js...")

    data = {
        "WEEKS": weeks,
        "WEEKS_ISO": weeks_iso,
        "WEEK_LABELS": week_labels,
        "SHOP_WEEKLY": shop_weekly,
        "WEEK_DATA": week_data,
        "TRAFFIC_WEEKLY": traffic_weekly,
        "PERSON_TARGETS": person_targets,
        "SKU_IMG": sku_img,
        "SKU_FIRST_DATE": merged_sku_first_date,
        "SKU_OWNER": sku_owner,
        "SKU_OWNER_LOOKUP": sku_owner_lookup,
        "SHOP_OWNERS": shop_owners,
        "SKU_WB_ID": sku_wb_id,
        "PRODUCT_NOTES": {},
        "NEW_PRODUCT_CREATED": [],
        "SKU_INVENTORY": sku_inventory,
        "NEW_PRODUCT_GRADE": new_product_grades,
        "SALES_GRADE": sales_grades,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    class SanitizedEncoder(json.JSONEncoder):
        def default(self, obj):
            if isinstance(obj, float):
                if math.isnan(obj) or math.isinf(obj):
                    return None
            return super().default(obj)

    json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'), cls=SanitizedEncoder)
    json_str = re.sub(r':NaN', ':null', json_str)
    json_str = re.sub(r':-Infinity', ':null', json_str)
    json_str = re.sub(r':Infinity', ':null', json_str)

    js_content = f'var DATA = {json_str};'

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(js_content)

    file_size = os.path.getsize(OUTPUT_PATH)
    print(f"  输出: {OUTPUT_PATH}")
    print(f"  大小: {file_size / 1024 / 1024:.2f} MB")

    print("\n验证 data_women.js 有效性...")
    try:
        if js_content.startswith('var DATA = ') and js_content.endswith(';'):
            inner = js_content[11:-1]
        else:
            inner = js_content
        verified = json.loads(inner)
        print(f"  OK DATA 有效: {len(verified)} 个顶级字段")
        for key in verified:
            v = verified[key]
            if isinstance(v, dict):
                print(f"    {key}: {len(v)} entries (dict)")
            elif isinstance(v, list):
                print(f"    {key}: {len(v)} entries (list)")
            else:
                print(f"    {key}: {type(v).__name__}")
    except json.JSONDecodeError as e:
        print(f"  FAIL DATA 无效: {e}")
        return 1

    print("\n" + "=" * 60)
    print("完成!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
