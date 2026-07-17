#!/usr/bin/env python3
"""
build_data.py - 从源Excel文件生成 data.json
源文件:
  - D:/周汇报文件/运营日数据.xlsx
  - D:/周汇报文件/产品列表.xlsx
  - D:/周汇报文件/LX利润表.xlsx
  - D:/周汇报文件/2026WB年规进度.xlsx
输出: output/data.json
"""

import json
import math
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl

# ── 配置 ──────────────────────────────────────────────
SRC_DIR = r"D:\周汇报文件"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "data.js")

# 俄语→中文类目映射
CATEGORY_CN = {
    "Шорты": "短裤", "Брюки": "长裤", "Брюки спортивные": "运动裤",
    "Джинсы": "牛仔裤", "Куртки": "夹克", "Пальто": "大衣",
    "Парки": "派克服", "Рубашки": "衬衫", "Свитеры": "毛衣",
    "Свитшоты": "卫衣", "Толстовки": "卫衣", "Футболки": "T恤",
    "Футболки-поло": "Polo衫", "Худи": "帽衫", "Юбки": "半身裙",
}

# ── 工具函数 ──────────────────────────────────────────

def sanitize_value(v):
    """确保值是合法JSON：NaN/Inf/-Inf → null，字符串数字 → float"""
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        # Python float → JSON number, round to 4 decimals for rates
        return round(v, 6) if abs(v) < 10 else round(v, 2)
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, str):
        # Excel 可能将数值读为字符串，尝试转换
        v_stripped = v.strip()
        if not v_stripped:
            return None
        try:
            f = float(v_stripped)
            if math.isnan(f) or math.isinf(f):
                return None
            return round(f, 6) if abs(f) < 10 else round(f, 2)
        except (ValueError, TypeError):
            return v
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return str(v)


def iso_week_to_date_range(iso_week_str):
    """'2025-W30' → (monday, sunday)"""
    year_s, week_s = iso_week_str.split("-W")
    year, week = int(year_s), int(week_s)
    jan4 = date(year, 1, 4)
    monday = jan4 - timedelta(days=jan4.weekday()) + timedelta(weeks=week - 1)
    return monday, monday + timedelta(days=6)


def date_to_iso_week(d):
    """date → '2025-W30'"""
    if isinstance(d, datetime):
        d = d.date()
    iso = d.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def load_workbook_safe(path):
    """安全加载workbook"""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return wb
    except Exception as e:
        print(f"ERROR: 无法加载 {path}: {e}")
        return None


# ── 阶段 1: 生成WEEKS等基础数据 ───────────────────────

def generate_weeks():
    """生成52周的数组"""
    weeks = []
    weeks_iso = []
    week_labels = []
    year, wn = 2025, 30
    for i in range(52):
        iso = f"{year}-W{wn:02d}"
        start, end = iso_week_to_date_range(iso)
        label = f"{iso} (W{i+1}·{start.month:02d}.{start.day:02d}-{end.month:02d}.{end.day:02d})"
        weeks.append(f"W{i+1}")
        weeks_iso.append(iso)
        week_labels.append(label)
        wn += 1
        if wn > 52:
            wn = 1
            year += 1
    return weeks, weeks_iso, week_labels


# ── 阶段 2: 读取产品列表 ──────────────────────────────

def read_product_list():
    """从产品列表.xlsx 提取 SKU映射"""
    path = os.path.join(SRC_DIR, "产品列表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None:
        return {}, {}, {}, {}, {}

    ws = wb[wb.sheetnames[0]]  # '产品列表'
    sku_img = {}
    sku_owner = {}
    sku_first_date = {}
    sku_wb_id = {}
    sku_category = {}
    shop_owner_set = defaultdict(set)

    # 列: 0=WB商品ID, 1=卖家SKU, ..., 9=主图, 10=店铺名称, 13=负责人, 14=创建时间
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if i > 2000:
            break  # safety limit
        vals = [cell.value for cell in row[:15]]
        wb_id = vals[0]
        sku = vals[1]
        img = vals[9]
        shop = vals[10] if len(vals) > 10 else None
        owner = vals[13] if len(vals) > 13 else None
        create_time = vals[14] if len(vals) > 14 else None
        cat_name = vals[6] if len(vals) > 6 else None

        if not sku or not wb_id:
            continue

        wb_id_str = str(int(wb_id)) if isinstance(wb_id, float) else str(wb_id)
        
        # 关键修改：使用 SKU+WB商品ID 组合作为唯一标识
        sku_wb_key = f"{sku}|{wb_id_str}"
        if shop:
            sku_wb_key_shop = f"{sku}|{shop}|{wb_id_str}"
        
        if img:
            sku_img[sku_wb_key] = str(img)
            if shop:
                sku_img[sku_wb_key_shop] = str(img)
        if owner:
            sku_owner[sku_wb_key] = str(owner)
            if shop:
                shop_owner_set[str(shop)].add(str(owner))
                sku_owner[sku_wb_key_shop] = str(owner)
        if create_time:
            fd_str = None
            if isinstance(create_time, (datetime, date)):
                fd_str = create_time.strftime("%Y-%m-%d") if isinstance(create_time, datetime) else create_time.isoformat()
            elif isinstance(create_time, str):
                fd_str = create_time[:10]
            if fd_str:
                sku_first_date[sku_wb_key] = fd_str
                if shop:
                    sku_first_date[sku_wb_key_shop] = fd_str
        
        # 存储 WB商品ID 映射
        sku_wb_id[sku_wb_key] = wb_id_str
        if shop:
            sku_wb_id[sku_wb_key_shop] = wb_id_str

        if cat_name:
            sku_category[str(sku)] = str(cat_name)

    # 转换 shop_owner_set → dict
    shop_owners = {s: {o: True for o in owners} for s, owners in shop_owner_set.items()}

    wb.close()
    print(f"  产品列表: {len(sku_img)} SKU图片, {len(sku_owner)} SKU负责人, "
          f"{len(shop_owners)} 店铺负责人, {len(sku_wb_id)} SKU-WB映射, {len(sku_first_date)} 首次日期, "
          f"{len(sku_category)} SKU类目")
    return sku_img, sku_owner, sku_first_date, sku_wb_id, shop_owners, sku_category


# ── 阶段 3: 读取LX利润表 ──────────────────────────────

def read_lx_profit(weeks_iso):
    """从LX利润表.xlsx 生成 WEEK_DATA 和 SHOP_WEEKLY"""
    path = os.path.join(SRC_DIR, "LX利润表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None:
        return {}, {}

    iso_to_w = {iso: f"W{i+1}" for i, iso in enumerate(weeks_iso)}

    # ─── 店铺分周利润表 → SHOP_WEEKLY ───
    ws_shop = wb["店铺分周利润表"]
    shop_weekly = defaultdict(dict)

    for i, row in enumerate(ws_shop.iter_rows(min_row=3)):
        if i > 5000:
            break
        vals = [cell.value for cell in row[:10]]
        week_end = vals[1]  # 星期结束值
        shop_name = vals[2]
        margin_val = vals[4]  # 毛利率CNY
        gsv_val = vals[6]    # 后台价GSV.CNY

        if not week_end or not shop_name:
            continue

        iso = date_to_iso_week(week_end) if isinstance(week_end, (datetime, date)) else str(week_end)
        w_key = iso_to_w.get(iso)
        if w_key is None:
            continue

        margin = sanitize_value(margin_val)
        gsv = sanitize_value(gsv_val)

        # 只保留 margin >= 0 的有效周数据
        if margin is not None and gsv is not None:
            # margin 在SHOP_WEEKLY中是百分比形式 (如 6.05 表示 6.05%)
            shop_weekly[str(shop_name)][w_key] = {
                "gsv": gsv,
                "margin": round(margin * 100, 4) if isinstance(margin, float) and margin < 1 else margin
            }

    # 转换为普通dict
    shop_weekly = {k: dict(v) for k, v in shop_weekly.items()}

    # ─── 分周SKU → WEEK_DATA ───
    ws_sku = wb["分周SKU"]
    # 列: 0=数据范围, 1=星期结束值, 2=店铺名称, 3=负责人, 4=子负责人, 5=类目,
    #     6=WB商品ID, 7=卖家SKU, 8=主图, 9=毛利量CNY, 10=毛利率CNY,
    #     11=GSV(后台价), 12=周订单量售完天数, 13=每周日库存量, 14=货值CNY,
    #     15=销售数量, 16=退款数量, 17=财报净销量, ..., 22=送达退货率,
    #     36=广告花费CNY

    week_data_raw = defaultdict(list)
    lx_owner_map = {}  # sku|shop → 子负责人（LX利润表 Col 4）

    sku_count = 0
    for i, row in enumerate(ws_sku.iter_rows(min_row=3)):
        if i > 40000:
            break
        vals = [cell.value for cell in row[:38]]
        week_end = vals[1]
        shop = vals[2]
        cat = vals[5]
        sku = vals[7]
        profit = vals[9]     # 毛利量CNY
        margin_rate = vals[10]  # 毛利率CNY (as decimal e.g., 0.2507)
        gsv = vals[11]       # GSV(后台价)
        qty = vals[15]       # 销售数量
        return_rate = vals[22] if len(vals) > 22 else None  # 送达退货率
        ad_spend = vals[36] if len(vals) > 36 else None  # 广告花费CNY
        unit_delivery = vals[18] if len(vals) > 18 else None  # 单件尾程配送CNY
        unit_return_fee = vals[19] if len(vals) > 19 else None  # 单件退货费CNY
        sub_owner = str(vals[4]).strip() if vals[4] else ""    # 子负责人

        if not week_end or not sku:
            continue

        iso = date_to_iso_week(week_end) if isinstance(week_end, (datetime, date)) else str(week_end)
        w_key = iso_to_w.get(iso)
        if w_key is None:
            continue

        sku_str = str(sku).strip() if sku else ""

        p_profit = sanitize_value(profit)
        p_gsv = sanitize_value(gsv)
        p_qty = sanitize_value(qty)
        p_ad = sanitize_value(ad_spend)
        p_del = sanitize_value(unit_delivery)
        p_ret = sanitize_value(unit_return_fee)

        product = {
            "sku": sku_str or "",
            "shop": str(shop) if shop else "",
            "cat": str(cat) if cat else "",
            "owner": sub_owner,
            "profit": p_profit if isinstance(p_profit, (int, float)) else 0,
            "margin": sanitize_value(margin_rate),
            "gsv": p_gsv if isinstance(p_gsv, (int, float)) else 0,
            "qty": p_qty if isinstance(p_qty, (int, float)) else 0,
            "return_rate": sanitize_value(return_rate),
            "ad_spend": p_ad if isinstance(p_ad, (int, float)) else 0,
            "unit_delivery": p_del if isinstance(p_del, (int, float)) else 0,
            "unit_return_fee": p_ret if isinstance(p_ret, (int, float)) else 0
        }

        # 子负责人映射（用于合并进 SKU_OWNER_LOOKUP）
        if sub_owner and sub_owner != "无负责人":
            lx_key = f"{sku_str}|{str(shop)}"
            if lx_key not in lx_owner_map:
                lx_owner_map[lx_key] = sub_owner

        # margin是小数, 转为百分比
        if isinstance(product["margin"], float) and product["margin"] < 1:
            product["margin"] = round(product["margin"] * 100, 2)
        # return_rate 也是小数, 同样转为百分比
        if isinstance(product["return_rate"], float) and product["return_rate"] < 1:
            product["return_rate"] = round(product["return_rate"] * 100, 2)

        week_data_raw[w_key].append(product)
        sku_count += 1

    # 构建最终的 WEEK_DATA
    week_data = {}
    for w_key in sorted(week_data_raw.keys(), key=lambda x: int(x[1:])):
        products = week_data_raw[w_key]

        # 按 profit 降序排列
        products.sort(key=lambda p: p["profit"], reverse=True)

        # 计算店铺汇总
        shop_summary = defaultdict(lambda: {"gsv": 0, "profit": 0, "margin": 0, "products": 0, "ad_spend": 0})
        for p in products:
            s = shop_summary[p["shop"]]
            s["gsv"] += p["gsv"]
            s["profit"] += p["profit"]
            s["products"] += 1
            s["ad_spend"] += p.get("ad_spend", 0) or 0

        # 计算 weighted margin
        for s in shop_summary.values():
            if s["gsv"] > 0:
                s["margin"] = round(s["profit"] / s["gsv"] * 100, 2)
            else:
                s["margin"] = 0

        # top10 by profit
        top10 = [p for p in products if p["sku"] != "无匹配ID费用"][:10]

        week_data[w_key] = {
            "shops": {k: dict(v) for k, v in shop_summary.items()},
            "top10Profit": top10,
            "allProducts": products
        }

    wb.close()
    print(f"  SHOP_WEEKLY: {len(shop_weekly)} shops")
    print(f"  WEEK_DATA: {len(week_data)} weeks, {sku_count} total product-weeks")
    return shop_weekly, week_data, lx_owner_map


# ── 阶段 4: 读取运营日数据 → TRAFFIC_WEEKLY ──────────

def read_traffic_weekly(weeks_iso):
    """从运营日数据.xlsx 生成 TRAFFIC_WEEKLY"""
    path = os.path.join(SRC_DIR, "运营日数据.xlsx")
    wb = load_workbook_safe(path)
    if wb is None:
        return {}

    iso_to_w = {iso: f"W{i+1}" for i, iso in enumerate(weeks_iso)}
    ws = wb[wb.sheetnames[0]]  # Export

    # 列: 0=日期(A), 1=主图CN(B), 2=店铺名称(C), 3=WB商品ID(D), 4=卖家SKU(E),
    #     5=可售数量(F), 8=访客(I), 11=加购数(L), 13=销量(N), 23=财报退货率(X), 27=广告点击率(AB)
    # 按 ISO周 + SKU 汇总: visitors, add_to_cart_count, sales_qty, click_cnt, return_qty, total_qty
    weekly_agg = defaultdict(lambda: defaultdict(lambda: {
        "visitors": 0, "atc": 0, "qty": 0, "gmv": 0,
        "click_cnt": 0, "click_impressions": 0,
        "return_qty": 0, "total_qty_ref": 0
    }))

    # 同时追踪每个SKU首次出现库存的日期（用于上架天数计算）
    sku_first_inventory_date = {}
    # 追踪每个SKU最新日期的可售数量（F列）
    sku_latest_inventory = {}  # sku -> {'date': date, 'value': float}

    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if i > 200000:
            break
        if not row:
            continue

        d_val = row[0].value if len(row) > 0 else None  # 日期 A列
        sku = row[4].value if len(row) > 4 else None     # 卖家SKU E列
        inventory = row[5].value if len(row) > 5 else None  # 可售数量 F列
        visitors = row[8].value if len(row) > 8 else None  # 访客 I列
        atc = row[11].value if len(row) > 11 else None     # 加购数 L列
        qty = row[13].value if len(row) > 13 else None     # 销量 N列
        gmv = row[14].value if len(row) > 14 else None      # GMV O列
        click_rate_val = row[27].value if len(row) > 27 else None  # 广告点击率 AB列
        return_rate_raw = row[23].value if len(row) > 23 else None  # 财报退货率 X列

        if not d_val or not sku:
            continue

        if isinstance(d_val, datetime):
            d_date = d_val.date()
        elif isinstance(d_val, date):
            d_date = d_val
        elif isinstance(d_val, str):
            try:
                d_date = datetime.strptime(d_val[:10], "%Y-%m-%d").date()
            except:
                continue
        else:
            continue

        iso = date_to_iso_week(d_date)
        w_key = iso_to_w.get(iso)
        if w_key is None:
            continue

        sku_str = str(sku).strip()
        agg = weekly_agg[w_key][sku_str]
        agg["visitors"] += float(visitors) if visitors else 0
        agg["atc"] += float(atc) if atc else 0
        agg["qty"] += float(qty) if qty else 0
        agg["gmv"] += float(gmv) if gmv else 0

        # 追踪每个SKU首次出现库存>0的日期（E列可售数量>0）
        if inventory is not None:
            try:
                inv_val = float(inventory)
                if inv_val > 0:
                    if sku_str not in sku_first_inventory_date or d_date < sku_first_inventory_date[sku_str]:
                        sku_first_inventory_date[sku_str] = d_date
                # 追踪最新日期的可售数量
                if sku_str not in sku_latest_inventory or d_date > sku_latest_inventory[sku_str]['date']:
                    sku_latest_inventory[sku_str] = {'date': d_date, 'value': inv_val}
            except (ValueError, TypeError):
                pass

        # 广告点击率: 累积每天的点击率值, 用于后续计算加权平均
        if click_rate_val is not None and visitors:
            click_rate_f = float(click_rate_val)
            if not (math.isnan(click_rate_f) or math.isinf(click_rate_f)):
                agg["click_cnt"] += click_rate_f * float(visitors)
                agg["click_impressions"] += float(visitors)

    # 转换为 TRAFFIC_WEEKLY 格式
    # [click_rate, add_to_cart_rate, conversion_rate, return_rate, sales_qty, gmv]
    traffic_weekly = {}
    for w_key, sku_data in sorted(weekly_agg.items(), key=lambda x: int(x[0][1:])):
        traffic_weekly[w_key] = {}
        for sku, agg in sku_data.items():
            v = agg["visitors"]
            # click_rate: 加权平均 (sum(click_rate * visitors) / sum(visitors))
            if agg["click_impressions"] > 0:
                click_rate = round(agg["click_cnt"] / agg["click_impressions"], 6)
            else:
                click_rate = None
            atc_rate = round(agg["atc"] / v, 6) if v > 0 else 0.0
            conv_rate = round(agg["qty"] / v, 6) if v > 0 else 0.0
            # 财报退货率暂不在 TRAFFIC_WEEKLY 中, 用 null 占位
            return_rate = None
            sales_qty = round(agg["qty"], 0)
            gmv_val = round(agg["gmv"], 2)
            traffic_weekly[w_key][sku] = [click_rate, atc_rate, conv_rate, return_rate, sales_qty, gmv_val]

    wb.close()
    total_entries = sum(len(v) for v in traffic_weekly.values())
    print(f"  TRAFFIC_WEEKLY: {len(traffic_weekly)} weeks, {total_entries} total SKU-week entries")
    print(f"  SKU首次库存日期 (运营日数据): {len(sku_first_inventory_date)} 个SKU")
    print(f"  SKU最新可售数量 (运营日数据): {len(sku_latest_inventory)} 个SKU")
    return traffic_weekly, sku_first_inventory_date, sku_latest_inventory


# ── 阶段 5: 读取年规进度 → PERSON_TARGETS ─────────────

def read_person_targets():
    """从2026WB年规进度.xlsx 生成 PERSON_TARGETS"""
    path = os.path.join(SRC_DIR, "2026WB年规进度.xlsx")
    wb = load_workbook_safe(path)
    if wb is None:
        return {}

    # 列映射: Excel列 → 月份
    # Col 2=2月, 3=3月, 4=4月, 5=5月, 6=6月, 7=7月, 8=8月,
    # 9=9月, 10=10月, 11=11月, 12=12月, 13=1月
    col_to_month = {2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7, 8: 8,
                    9: 9, 10: 10, 11: 11, 12: 12, 13: 1}

    # 行到字段的映射（不同sheet可能有不同结构）
    # 我们需要从列B识别
    all_targets = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        person_name = sheet_name.strip()

        # 读取所有行
        rows_data = {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100)):
            vals = [cell.value for cell in row[:16]]
            rows_data[i+1] = vals

        # 从B列识别指标
        targets_by_month = defaultdict(dict)

        shop_section_started = False
        for row_num, vals in rows_data.items():
            # 遇到第一个店铺名行后，停止处理（后续都是逐店明细）
            a_val = vals[0] if len(vals) > 0 else None
            if a_val and isinstance(a_val, str) and "店" in a_val:
                shop_section_started = True
                continue
            if shop_section_started:
                continue  # 已进入逐店排分区，跳过所有后续行
            label = str(vals[1]).strip() if vals[1] is not None else ""
            if not label:
                continue

            # 月度目标 利润 → profit_target
            if "月度目标" in label and "利润" in label:
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["profit_target"] = sanitize_value(v) or 0

            # 实际利润 → profit_done（排除 "利润率" 行）
            elif "实际利润" in label and "率" not in label:
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["profit_done"] = sanitize_value(v) or 0

            # 销量目标 → sales_target（排除 "销量完成进度"）
            elif label == "销量目标":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["sales_target"] = sanitize_value(v) or 0

            # 销量完成 → sales_done（排除 "销量完成进度"）
            elif label == "销量完成":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["sales_done"] = sanitize_value(v) or 0

            # GMV目标
            elif label == "GMV目标":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["gmv_target"] = sanitize_value(v) or 0

            # GMV完成
            elif label == "GMV完成":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["gmv_done"] = sanitize_value(v) or 0

            # GSV目标
            elif label == "GSV目标":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["gsv_target"] = sanitize_value(v) or 0

            # GSV完成
            elif label == "GSV完成":
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["gsv_done"] = sanitize_value(v) or 0

            # 利润率（直接读取 Excel 行，优先于代码计算）
            elif "利润率" in label and "率" in label:
                for col, month in col_to_month.items():
                    v = vals[col]
                    if v is not None and month >= 1 and month <= 12:
                        targets_by_month[month]["profit_rate_excel"] = sanitize_value(v)

        # 计算 profit_rate（Excel 行优先，无则 fallback 到 profit_done/gsv_done）
        for month in targets_by_month:
            t = targets_by_month[month]
            excel_rate = t.get("profit_rate_excel")
            if excel_rate is not None and isinstance(excel_rate, (int, float)) and excel_rate > 0:
                # Excel 已存为小数（如 0.1319），转换为百分比
                t["profit_rate"] = round(excel_rate * 100, 2)
            else:
                gsv_done = t.get("gsv_done", 0)
                profit_done = t.get("profit_done", 0)
                if gsv_done and gsv_done > 0 and profit_done:
                    t["profit_rate"] = round(profit_done / gsv_done * 100, 2)
                else:
                    t["profit_rate"] = 0

        # 确保所有12个月都有数据
        for m in range(1, 13):
            if m not in targets_by_month:
                targets_by_month[m] = {
                    "gsv_target": 0, "gsv_done": 0,
                    "gmv_target": 0, "gmv_done": 0,
                    "sales_target": 0, "sales_done": 0,
                    "profit_target": 0, "profit_done": 0,
                    "profit_rate": 0
                }

        # 转换 key 为字符串
        person_data = {str(m): dict(targets_by_month[m]) for m in range(1, 13)}
        all_targets[person_name] = person_data

    wb.close()
    print(f"  PERSON_TARGETS: {len(all_targets)} people")
    return all_targets


# ── 阶段 5.5: 全量预计算新品等级和销售等级 ──────────────

def compute_new_product_grades(week_data, sku_first_date, weeks_iso):
    """全量预计算新品等级（45天规则，按销量）：
    对每个 SKU，取首次库存日期后 45 天内累计 qty（销量）。
    阈值：S≥175, A≥84, B≥56。不满足不打标签。
    """
    all_products = set()
    for wk, wd in week_data.items():
        for p in wd.get("allProducts", []):
            shop = p.get("shop", "")
            sku = p.get("sku", "")
            if shop and sku and sku != "无匹配ID费用":
                all_products.add(f"{shop}|{sku}")

    iso_to_dates = {}
    for iso in weeks_iso:
        start, end = iso_week_to_date_range(iso)
        iso_to_dates[iso] = (start, end)

    # iso → W-key 映射（修复：原 w_to_iso 方向反了）
    iso_to_w = {iso: f"W{i+1}" for i, iso in enumerate(weeks_iso)}

    today = date.today()
    new_product_grades = {}

    for key in all_products:
        parts = key.split("|", 1)
        shop = parts[0]
        sku = parts[1] if len(parts) > 1 else ""

        fd = None
        for lk in [key, f"{sku}|{shop}", sku]:
            if lk in sku_first_date:
                fd = sku_first_date[lk]
                break

        if not fd:
            new_product_grades[key] = None
            continue

        try:
            parts_fd = fd.split("-")
            first_date = date(int(parts_fd[0]), int(parts_fd[1]), int(parts_fd[2]))
        except (ValueError, IndexError):
            new_product_grades[key] = None
            continue

        cutoff_date = first_date + timedelta(days=45)

        if first_date > today:
            new_product_grades[key] = None
            continue

        # 累加45天窗口内的销量（qty）
        total_qty = 0.0
        for iso in weeks_iso:
            ws, we = iso_to_dates[iso]
            if we >= first_date and ws <= cutoff_date:
                w_key = iso_to_w.get(iso)
                if w_key and w_key in week_data:
                    for p in week_data[w_key].get("allProducts", []):
                        if p.get("shop") == shop and p.get("sku") == sku:
                            total_qty += p.get("qty", 0) or 0

        # 按销量阈值：S≥175, A≥84, B≥56
        if total_qty >= 175:
            new_product_grades[key] = "S"
        elif total_qty >= 84:
            new_product_grades[key] = "A"
        elif total_qty >= 56:
            new_product_grades[key] = "B"
        else:
            new_product_grades[key] = None

    s_count = sum(1 for v in new_product_grades.values() if v == "S")
    a_count = sum(1 for v in new_product_grades.values() if v == "A")
    b_count = sum(1 for v in new_product_grades.values() if v == "B")
    print(f"  NEW_PRODUCT_GRADE (按销量): {len(new_product_grades)} SKUs (S={s_count}, A={a_count}, B={b_count})")
    return new_product_grades


def compute_sales_grades(week_data, sku_first_date, weeks_iso, sku_inventory):
    """全量预计算销售等级（历史最高周，按销量）：
    对每个 SKU，遍历所有历史周找出单周最高销量。
    阈值：S≥350, A≥175, B≥105。不满足不打标签。
    """
    all_products = set()
    for wk, wd in week_data.items():
        for p in wd.get("allProducts", []):
            shop = p.get("shop", "")
            sku = p.get("sku", "")
            if shop and sku and sku != "无匹配ID费用":
                all_products.add(f"{shop}|{sku}")

    sales_grades = {}

    for key in all_products:
        parts = key.split("|", 1)
        shop = parts[0]
        sku = parts[1] if len(parts) > 1 else ""

        # 遍历所有周，找单周最高销量
        max_weekly_qty = 0.0
        for wk, wd in week_data.items():
            for p in wd.get("allProducts", []):
                if p.get("shop") == shop and p.get("sku") == sku:
                    qty = p.get("qty", 0) or 0
                    if qty > max_weekly_qty:
                        max_weekly_qty = qty

        # 按单周最高销量阈值：S≥350, A≥175, B≥105
        if max_weekly_qty >= 350:
            sales_grades[key] = "S"
        elif max_weekly_qty >= 175:
            sales_grades[key] = "A"
        elif max_weekly_qty >= 105:
            sales_grades[key] = "B"
        else:
            sales_grades[key] = None

    s_count = sum(1 for v in sales_grades.values() if v == "S")
    a_count = sum(1 for v in sales_grades.values() if v == "A")
    b_count = sum(1 for v in sales_grades.values() if v == "B")
    none_count = sum(1 for v in sales_grades.values() if v is None)
    print(f"  SALES_GRADE (单周最高销量): {len(sales_grades)} SKUs (S={s_count}, A={a_count}, B={b_count}, None={none_count})")
    return sales_grades


# ── 主函数 ────────────────────────────────────────────

def main():
    print("=" * 60)
    print("build_data.py - 生成 data.json")
    print("=" * 60)

    # 阶段 1: 基础周数组
    print("\n[1/5] 生成周数组...")
    weeks, weeks_iso, week_labels = generate_weeks()
    print(f"  共 {len(weeks)} 周: {weeks_iso[0]} → {weeks_iso[-1]}")

    # 阶段 1.5: 计算 MONTH_WEEK_MAP（total 按日历维度计算）
    month_week_map = {}
    from datetime import date as dt_date, timedelta as dt_timedelta
    for iso, w in zip(weeks_iso, weeks):
        iso_year, iso_wn = int(iso[:4]), int(iso[6:])
        monday = dt_date.fromisocalendar(iso_year, iso_wn, 1)
        sunday = monday + dt_timedelta(days=6)
        month_key = f"{sunday.year:04d}-{sunday.month:02d}"
        if month_key not in month_week_map:
            month_week_map[month_key] = {"weeks": [], "total": 0}
        month_week_map[month_key]["weeks"].append(w)
    # 按日历维度计算每个月实际包含的 ISO 周数（total）
    for month_key in month_week_map:
        y, m = int(month_key[:4]), int(month_key[5:])
        month_start = dt_date(y, m, 1)
        month_end = dt_date(y, m + 1, 1) - dt_timedelta(days=1) if m < 12 else dt_date(y, 12, 31)
        iso_weeks_in_month = set()
        d = month_start
        while d <= month_end:
            iso_weeks_in_month.add(d.isocalendar()[1])
            d += dt_timedelta(days=1)
        month_week_map[month_key]["total"] = len(iso_weeks_in_month)
    print(f"  MONTH_WEEK_MAP: {len(month_week_map)} months")

    # 阶段 2: 产品列表
    print("\n[2/5] 读取产品列表...")
    sku_img, sku_owner, sku_first_date, sku_wb_id, shop_owners, sku_category = read_product_list()

    # 阶段 3: 利润表
    print("\n[3/5] 读取LX利润表...")
    shop_weekly, week_data, lx_owner_map = read_lx_profit(weeks_iso)

    # 阶段 4: 运营日数据
    print("\n[4/5] 读取运营日数据...")
    traffic_weekly, sku_first_inventory_date, sku_latest_inventory = read_traffic_weekly(weeks_iso)

    # 阶段 4.5: 将运营日数据 M列(销量) 合并到 WEEK_DATA.qty
    print("  合并运营日数据 M列(销量) 到 WEEK_DATA...")
    qty_merged = 0
    for w_key in week_data:
        if w_key not in traffic_weekly:
            continue
        sku_qty = traffic_weekly[w_key]  # dict: sku -> [click_rate, atc_rate, conv_rate, return_rate, sales_qty, gmv]
        for p in week_data[w_key]["allProducts"]:
            sku = p["sku"]
            if sku in sku_qty and len(sku_qty[sku]) > 4:
                new_qty = sku_qty[sku][4]  # index 4 = sales_qty from 运营日数据 M列
                if new_qty and new_qty > 0:
                    p["qty"] = new_qty
                    qty_merged += 1
                gmv_val = sku_qty[sku][5] if len(sku_qty[sku]) > 5 else 0
                p["gmv"] = gmv_val if gmv_val else 0
    print(f"  已合并 {qty_merged} 条销量数据 (来源: 运营日数据 M列)")

    # 构建 SKU 到 WB_IDs 的映射
    sku_to_wb_ids = {}
    for key in sku_wb_id:
        parts = key.split('|')
        if len(parts) >= 2:
            sku = parts[0]
            wb_id = sku_wb_id[key]
            if sku not in sku_to_wb_ids:
                sku_to_wb_ids[sku] = set()
            sku_to_wb_ids[sku].add(wb_id)
    
    # SKU_FIRST_DATE: 合并两个来源
    # 1. 产品列表中的创建时间（已有 SKU+WB_ID 组合 key）
    # 2. 运营日数据首次库存>0的日期（仅 SKU key，作为补充/优先数据）
    merged_sku_first_date = {}
    
    # 先从运营日数据获取首次库存>0日期（优先级更高，因为反映真实上架）
    for sku, d in sku_first_inventory_date.items():
        date_str = d.strftime("%Y-%m-%d") if isinstance(d, date) else str(d)[:10]
        # 仅 SKU 级别 key
        merged_sku_first_date[sku] = date_str
        # 同时为每个 WB_ID 变体设置相同日期（因为运营日数据没有 WB_ID 区分）
        if sku in sku_to_wb_ids:
            for wb_id in sku_to_wb_ids[sku]:
                merged_sku_first_date[f"{sku}|{wb_id}"] = date_str
    
    # 再从产品列表补充（仅补充运营日数据中没有的）
    for key, date_str in sku_first_date.items():
        if key not in merged_sku_first_date:
            merged_sku_first_date[key] = date_str
    
    # 补充 shop|sku 格式的 key，供 compute_new_product_grades 使用
    # compute_new_product_grades 的 all_products key 格式为 shop|sku
    shop_sku_added = 0
    for wk, wd in week_data.items():
        for p in wd.get("allProducts", []):
            shop = p.get("shop", "")
            sku = p.get("sku", "")
            if not shop or not sku or sku == "无匹配ID费用":
                continue
            shop_sku_key = f"{shop}|{sku}"
            if shop_sku_key in merged_sku_first_date:
                continue
            # 尝试从已有 key 中匹配日期
            fd = None
            for candidate in [sku, f"{sku}|{shop}"]:
                if candidate in merged_sku_first_date:
                    fd = merged_sku_first_date[candidate]
                    break
            if not fd and sku in sku_to_wb_ids:
                for wb_id in sku_to_wb_ids[sku]:
                    for candidate in [f"{sku}|{wb_id}", f"{sku}|{shop}|{wb_id}"]:
                        if candidate in merged_sku_first_date:
                            fd = merged_sku_first_date[candidate]
                            break
                    if fd:
                        break
            if fd:
                merged_sku_first_date[shop_sku_key] = fd
                shop_sku_added += 1

    print(f"  SKU_FIRST_DATE: {len(merged_sku_first_date)} 个 (合并产品列表+运营日数据), 新增 shop|sku 映射 {shop_sku_added} 个")
    print(f"    无库存记录的SKU将不显示上架天数（显示为 '-'）")

    # SKU_INVENTORY: 每个SKU最新日期的可售数量
    sku_inventory = {}
    for sku, info in sku_latest_inventory.items():
        sku_inventory[sku] = int(info['value'])

    # 添加 shop|sku 映射，让渲染代码能通过 SKU代码|店铺 键查找库存
    inv_shop_sku_added = 0
    for wk, wd in week_data.items():
        for p in wd.get("allProducts", []):
            shop = p.get("shop", "")
            sku = p.get("sku", "")
            if not shop or not sku or sku == "无匹配ID费用":
                continue
            shop_sku_key = f"{shop}|{sku}"
            if shop_sku_key in sku_inventory:
                continue
            if sku in sku_to_wb_ids:
                for wb_id in sku_to_wb_ids[sku]:
                    wb_id_str = str(wb_id)
                    if wb_id_str in sku_inventory:
                        sku_inventory[shop_sku_key] = sku_inventory[wb_id_str]
                        inv_shop_sku_added += 1
                        break

    print(f"  SKU_INVENTORY: {len(sku_inventory)} 个SKU (最新日期可售数量), 新增 shop|sku 映射 {inv_shop_sku_added} 个")

    # 阶段 5: 年规进度
    print("\n[5/5] 读取年规进度...")
    person_targets = read_person_targets()

    # ── 构建 SKU_OWNER_LOOKUP 映射表 ──
    # 将 SKU_OWNER 的原始 key（sku|shop|wb_id 或 sku|wb_id）映射为
    # JS 端可直接查询的扁平格式：sku|shop → owner, sku → owner
    sku_owner_lookup = {}
    for key, owner in sku_owner.items():
        parts = key.split('|')
        # 3-part key: sku|shop_name|id → map as sku|shop_name → owner
        if len(parts) == 3:
            lookup_key = f"{parts[0]}|{parts[1]}"
            sku_owner_lookup[lookup_key] = owner
        # 2-part key: sku|id → map as sku → owner (only if no shop-level mapping yet)
        if parts[0] not in sku_owner_lookup:
            sku_owner_lookup[parts[0]] = owner
    print(f"  SKU_OWNER_LOOKUP: {len(sku_owner_lookup)} entries (before LX merge)")

    # ── 合并 LX利润表子负责人到 SKU_OWNER_LOOKUP ──
    lx_merged = 0
    for key, sub_owner in lx_owner_map.items():
        if sub_owner:
            sku_owner_lookup[key] = sub_owner  # 子负责人覆盖原负责人
            lx_merged += 1
            # 同时为纯 sku key 设置（仅当不存在时）
            sku_only = key.split('|')[0]
            if sku_only not in sku_owner_lookup:
                sku_owner_lookup[sku_only] = sub_owner
    print(f"  SKU_OWNER_LOOKUP: {len(sku_owner_lookup)} entries (after LX merge, +{lx_merged})")

    # ── 更新 shop_owners：只把LX子负责人加到他们实际有产品的店铺 ──
    for key, sub_owner in lx_owner_map.items():
        parts = key.split('|')
        if len(parts) >= 2 and sub_owner:
            shop = parts[1]
            if shop not in shop_owners:
                shop_owners[shop] = {}
            if sub_owner not in shop_owners[shop]:
                shop_owners[shop][sub_owner] = True
    print(f"  SHOP_OWNERS updated with LX sub-owners (per-shop accurate)")

    # ── 阶段 5.5: 预计算新品等级和销售等级 ──
    print("\n[5.5] 预计算新品等级和销售等级...")
    new_product_grades = compute_new_product_grades(
        week_data, merged_sku_first_date, weeks_iso
    )
    sales_grades = compute_sales_grades(
        week_data, merged_sku_first_date, weeks_iso, sku_inventory
    )

    # 类目中文化
    sku_category_cn = {}
    for sku, cat in sku_category.items():
        sku_category_cn[sku] = CATEGORY_CN.get(cat, cat)
    print(f"  SKU_CATEGORY 中文化: {len(sku_category_cn)} entries")

    # ── 组装 data.json ──
    print("\n" + "=" * 60)
    print("组装 data.json...")

    data = {
        "WEEKS": weeks,
        "WEEKS_ISO": weeks_iso,
        "WEEK_LABELS": week_labels,
        "SHOP_WEEKLY": shop_weekly,
        "WEEK_DATA": week_data,
        "TRAFFIC_WEEKLY": traffic_weekly,
        "PERSON_TARGETS": person_targets,
        "SKU_IMG": sku_img,
        "SKU_FIRST_DATE": merged_sku_first_date,
        "SKU_OWNER": sku_owner,
        "SKU_OWNER_LOOKUP": sku_owner_lookup,
        "SHOP_OWNERS": shop_owners,
        "SKU_WB_ID": sku_wb_id,
        "PRODUCT_NOTES": {},
        "NEW_PRODUCT_CREATED": [],
        "SKU_INVENTORY": sku_inventory,
        "NEW_PRODUCT_GRADE": new_product_grades,
        "SALES_GRADE": sales_grades,
        "SKU_CATEGORY": sku_category_cn,
        "MONTH_WEEK_MAP": month_week_map,
        "OWNERS": sorted(set(o for owners in shop_owners.values() for o in owners) | set(v for v in lx_owner_map.values() if v)) + ["其他/待定"],
        "OWNER_HIERARCHY": {
            "江凯伦": ["林梓蕾", "陈欣诺"],
            "张梦瑶": ["何欢洁", "郑志远"]
        },
    }

    # ── 写入 data.js ──
    # 确保 output 目录存在
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    # 自定义 JSON encoder 处理特殊值
    class SanitizedEncoder(json.JSONEncoder):
        def default(self, obj):
            if isinstance(obj, float):
                if math.isnan(obj) or math.isinf(obj):
                    return None
            return super().default(obj)

    json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'), cls=SanitizedEncoder)
    # 二次检查：确保没有任何 NaN/Infinity 出现在 JSON 中
    json_str = re.sub(r':NaN', ':null', json_str)
    json_str = re.sub(r':-Infinity', ':null', json_str)
    json_str = re.sub(r':Infinity', ':null', json_str)

    # 包装为 JS 变量声明（通过 <script src="data.js"> 加载，无需 fetch/XHR）
    js_content = f'var DATA = {json_str};'

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(js_content)

    file_size = os.path.getsize(OUTPUT_PATH)
    print(f"  输出: {OUTPUT_PATH}")
    print(f"  大小: {file_size / 1024 / 1024:.2f} MB")

    # ── 验证 data.js 有效性 ──
    print("\n验证 data.js 有效性...")
    try:
        # 提取 JSON 部分（去除 var DATA =  和尾部 ;）
        if js_content.startswith('var DATA = ') and js_content.endswith(';'):
            inner = js_content[11:-1]
        else:
            inner = js_content
        verified = json.loads(inner)
        print(f"  ✓ DATA 有效: {len(verified)} 个顶级字段")
        for key in verified:
            v = verified[key]
            if isinstance(v, dict):
                print(f"    {key}: {len(v)} entries (dict)")
            elif isinstance(v, list):
                print(f"    {key}: {len(v)} entries (list)")
            else:
                print(f"    {key}: {type(v).__name__}")
    except json.JSONDecodeError as e:
        print(f"  ❌ DATA 无效: {e}")
        return 1

    # ── 自动更新 HTML 中 data.js 的版本号（强制浏览器刷新缓存）──
    html_path = os.path.join(OUTPUT_DIR, "product-weekly-report.html")
    if os.path.exists(html_path):
        new_ver = datetime.now().strftime("%Y%m%d%H%M")
        with open(html_path, 'r', encoding='utf-8') as f:
            html = f.read()
        updated_html = re.sub(r'data\.js\?v=\d+', f'data.js?v={new_ver}', html)
        if updated_html != html:
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(updated_html)
            print(f"\n  已更新 HTML 缓存版本号: ?v={new_ver}")
        else:
            print(f"\n  ⚠ HTML 中未找到 data.js?v= 模式，请手动检查")

    # ── 数据质量诊断 ──
    print("\n数据质量诊断:")
    # 检查 return_rate 是否为百分比
    rr_check = 0
    rr_small = 0
    qty_from_m = 0
    for wk in verified["WEEK_DATA"]:
        for p in verified["WEEK_DATA"][wk].get("allProducts", []):
            rr = p.get("return_rate", 0)
            if rr and rr > 0:
                rr_check += 1
                if rr < 1:
                    rr_small += 1
            if p.get("qty", 0) > 0:
                qty_from_m += 1
    print(f"  return_rate 百分比格式: {rr_check - rr_small}/{rr_check} (小数值 {rr_small})")
    print(f"  qty (来源 M列): {qty_from_m} 个产品")
    print(f"  TRAFFIC_WEEKLY: {len(verified['TRAFFIC_WEEKLY'])} 周")
    print(f"  SKU_OWNER: {len(verified['SKU_OWNER'])} 个 SKU")

    print("\n" + "=" * 60)
    print("完成!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
