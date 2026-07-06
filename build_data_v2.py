"""
build_data_v2.py - 男装版重构
从 LX利润表 A列动态取周，从年规读取目标
计算: 利润/GSV/实销/GMV/利润率/退款率/完成进度
"""
import json, os, math
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook

SRC_DIR = r"D:\周汇报文件"

def load_workbook_safe(path):
    try:
        return load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[WARN] Cannot open {path}: {e}")
        return None

def parse_date_str(s):
    """解析 24.09.02 或 2024.09.02 → datetime"""
    parts = s.split('.')
    if len(parts) != 3: return None
    y, m, d = parts
    if len(y) == 2: y = '20' + y
    try:
        return datetime(int(y), int(m), int(d))
    except:
        return None

# ============================================================
# 阶段 1: 从 LX利润表 A列动态生成周数组
# ============================================================

def generate_weeks_from_profit():
    path = os.path.join(SRC_DIR, "LX利润表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None:
        raise FileNotFoundError(f"无法打开 {path}")
    ws = wb["分周SKU"]

    week_ranges = set()
    week_end_dates = {}
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        dr, ed = row[0], row[1]
        if dr and isinstance(dr, str) and '-' in dr:
            week_ranges.add(dr)
            if ed and isinstance(ed, datetime):
                week_end_dates[dr] = ed

    sorted_ranges = sorted(week_ranges, key=lambda x: week_end_dates.get(x, datetime(2000,1,1)))

    weeks, weeks_iso, week_labels = [], [], []
    week_years, week_months, week_year_months = [], [], []

    for i, dr in enumerate(sorted_ranges):
        wk = f"W{i+1}"
        weeks.append(wk)
        parts = dr.split('-')
        year_num, month_num = 0, 0
        if len(parts) == 2:
            end_dt = parse_date_str(parts[1])
            if end_dt:
                year_num, month_num = end_dt.year, end_dt.month
                weeks_iso.append(f"{end_dt.year}-W{end_dt.isocalendar()[1]:02d}")
                week_labels.append(f"{year_num}年{month_num:02d}月 W{i+1}·{dr}")
            else:
                weeks_iso.append(f"WK{i+1}")
                week_labels.append(f"{dr} (W{i+1})")
        else:
            weeks_iso.append(f"WK{i+1}")
            week_labels.append(f"{dr} (W{i+1})")
        week_years.append(year_num)
        week_months.append(month_num)
        week_year_months.append(f"{year_num}-{month_num:02d}" if year_num else "")

    print(f"  读取 {len(weeks)} 周: {sorted_ranges[0]} → {sorted_ranges[-1]}")
    ym_set = sorted(set(w for w in week_year_months if w))
    print(f"  年月跨度: {ym_set[0]} → {ym_set[-1]} ({len(ym_set)} 个月)")
    return weeks, weeks_iso, week_labels, sorted_ranges, week_years, week_months, week_year_months

# ============================================================
# 阶段 2: 读取产品列表（含 shop→person 映射）
# ============================================================

def read_product_list():
    path = os.path.join(SRC_DIR, "产品列表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None: return {}, {}, {}, {}

    ws = wb[wb.sheetnames[0]]
    sku_img, sku_owner, sku_first_date, sku_wb_id = {}, {}, {}, {}
    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row[:15]]
        wb_id, sku, img, cat, shop = vals[0], vals[1], vals[4], vals[6], vals[11]
        owner = vals[9] or ""  # 负责人
        if not sku: continue
        sku = str(sku).strip()
        sku_wb_id[sku] = str(wb_id) if wb_id else ""
        sku_img[sku] = str(img) if img else ""
        sku_owner[sku] = str(owner) if owner else ""
        if vals[2]: sku_first_date[sku] = str(vals[2])[:10]

    print(f"  产品列表: {len(sku_img)} SKU, {len(sku_owner)} 负责人")
    return sku_img, sku_owner, sku_first_date, sku_wb_id

# ============================================================
# 阶段 3: 年规目标
# ============================================================

def read_targets():
    path = os.path.join(SRC_DIR, "2026WB年规进度.xlsx")
    wb = load_workbook_safe(path)
    if wb is None: return {}, {}
    person_targets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 第2行：月份
        month_cols = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if val and isinstance(val, str):
                for m_name in ["2月","3月","4月","5月","6月","7月","8月","9月","10月","11月","12月","1月"]:
                    if m_name in val:
                        month_cols[col] = int(m_name.replace("月",""))
                        break
        if not month_cols:
            continue

        target_rows = {
            "profit_target": 6, "sales_target": 10,
            "real_sales_target": 14, "gmv_target": 18, "gsv_target": 22,
        }
        person_data = {}
        for key, row_num in target_rows.items():
            for col, month in month_cols.items():
                val = ws.cell(row=row_num, column=col).value
                try: val = float(val) if val is not None else 0
                except: val = 0
                m_key = str(month)
                if m_key not in person_data: person_data[m_key] = {}
                person_data[m_key][key] = val

        # 固定目标
        for m_key in person_data:
            person_data[m_key]["margin_target"] = 15.0
            person_data[m_key]["return_target"] = 65.0

        person_targets[sheet_name] = person_data

    print(f"  年规目标: {len(person_targets)} 人")
    return person_targets

# ============================================================
# 阶段 4: 利润表实际数据 + 利润率/退款率
# ============================================================

def read_profit_data(week_ranges):
    path = os.path.join(SRC_DIR, "LX利润表.xlsx")
    wb = load_workbook_safe(path)
    if wb is None: return {}, {}, {}, {}
    ws = wb["分周SKU"]

    shop_weekly = defaultdict(lambda: defaultdict(lambda: {"gsv": 0, "profit": 0, "qty": 0, "real_sales": 0, "products": 0}))
    week_data = defaultdict(lambda: {"shops": {}, "allProducts": []})
    person_data = defaultdict(lambda: defaultdict(lambda: {"gsv": 0, "profit": 0, "qty": 0, "real_sales": 0}))

    total_rows = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 21: continue
        dr = row[0]
        if not dr or not isinstance(dr, str): continue
        shop = str(row[2] or "")
        owner = str(row[3] or "")
        sku = str(row[6] or "")
        if not shop or not sku: continue
        profit = float(row[8] or 0)       # I
        gsv = float(row[10] or 0)          # K
        real_sales = float(row[16] or 0)   # Q
        qty = int(row[19] or 0)            # T

        total_rows += 1
        shop_weekly[shop][dr]["gsv"] += gsv
        shop_weekly[shop][dr]["profit"] += profit
        shop_weekly[shop][dr]["real_sales"] += real_sales
        shop_weekly[shop][dr]["qty"] += qty
        shop_weekly[shop][dr]["products"] += 1

        if shop not in week_data[dr]["shops"]:
            week_data[dr]["shops"][shop] = {"gsv": 0, "profit": 0, "margin": 0, "products": 0, "ad_spend": 0}
        sd = week_data[dr]["shops"][shop]
        sd["gsv"] += gsv; sd["profit"] += profit; sd["products"] += 1

        week_data[dr]["allProducts"].append({
            "sku": sku, "shop": shop,
            "profit": round(profit, 2), "gsv": round(gsv, 2),
            "qty": qty, "real_sales": round(real_sales, 2),
        })

        # Person-level aggregation
        if owner:
            person_data[owner][dr]["gsv"] += gsv
            person_data[owner][dr]["profit"] += profit
            person_data[owner][dr]["real_sales"] += real_sales
            person_data[owner][dr]["qty"] += qty

    # 计算 margin
    for shop, weeks in shop_weekly.items():
        for wk, d in weeks.items():
            d["margin"] = round(d["profit"] / d["gsv"] * 100, 2) if d["gsv"] > 0 else 0
    for wk, wd in week_data.items():
        for shop, sd in wd["shops"].items():
            sd["margin"] = round(sd["profit"] / sd["gsv"] * 100, 2) if sd["gsv"] > 0 else 0
        wd["top10Profit"] = sorted(wd["allProducts"], key=lambda x: -x["profit"])[:10]

    print(f"  利润表: {len(shop_weekly)} shops, {len(week_data)} weeks, {total_rows} rows, {len(person_data)} persons")
    return dict(shop_weekly), dict(week_data), dict(person_data)

# ============================================================
# 阶段 5: 运营日数据（GMV）
# ============================================================

def read_traffic_and_gmv(week_ranges):
    path = os.path.join(SRC_DIR, "运营日数据.xlsx")
    wb = load_workbook_safe(path)
    if wb is None: return defaultdict(dict), {}
    ws = wb[wb.sheetnames[0]]

    gmv_data = defaultdict(lambda: defaultdict(float))  # week -> shop -> gmv
    gmv_person = defaultdict(lambda: defaultdict(float))  # week -> person -> gmv
    shop_owner_map = {}
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 16: continue
        date_val = row[0]
        shop = str(row[2] or "")
        gmv_val = float(row[14] or 0)
        if not shop or not date_val: continue

        # 匹配周范围
        matched = None
        if isinstance(date_val, datetime):
            for dr in week_ranges:
                parts = dr.split('-')
                if len(parts) != 2: continue
                sd = parse_date_str(parts[0])
                ed = parse_date_str(parts[1])
                if sd and ed and sd <= date_val <= ed:
                    matched = dr; break
        if matched:
            gmv_data[matched][shop] += gmv_val
            total += 1

    # Build person-level GMV using shop->person mapping from profit table
    # First, build shop -> primary_owner mapping
    shop_owner_map = {}
    path2 = os.path.join(SRC_DIR, "LX利润表.xlsx")
    wb2 = load_workbook_safe(path2)
    if wb2:
        ws_shop = wb2["店铺分周利润表"]
        # Build shop->owner from 负责人分周利润表
        ws_owner = wb2["负责人分周利润表"]
        for row in ws_owner.iter_rows(min_row=3, values_only=True):
            if row and len(row) >= 4:
                shop_name = str(row[2] or "")
                owner_name = str(row[3] or "")
                if shop_name and owner_name:
                    shop_owner_map[shop_name] = owner_name

    # Aggregate person-level GMV from traffic data
    gmv_person = defaultdict(lambda: defaultdict(float))
    for dr, shops in gmv_data.items():
        for shop, gmv_val in shops.items():
            owner = shop_owner_map.get(shop, "")
            if owner:
                gmv_person[dr][owner] += gmv_val

    traffic_weekly = {}
    for dr, shops in gmv_data.items():
        traffic_weekly[dr] = {}
        for shop, gmv in shops.items():
            traffic_weekly[dr][shop] = {
                "gmv": round(gmv, 2),
                "visitors": 0, "atc": 0, "qty": 0,
                "click_rate": 0, "cart_rate": 0, "conv_rate": 0,
            }

    print(f"  GMV: {total} entries, {len(gmv_data)} weeks")
    return traffic_weekly, dict(gmv_person)

# ============================================================
# 阶段 6: 月份归属 + 完成进度计算
# ============================================================

def month_of_range(dr):
    """取周结束日期的月份"""
    parts = dr.split('-')
    if len(parts) == 2:
        dt = parse_date_str(parts[1])
        if dt: return dt.month
    return None

def compute_monthly(person_data, gmv_person, person_targets, week_ranges, week_months):
    """按月汇总实际值，计算完成进度"""
    Monthly = type('Monthly', (), {})()
    Monthly.actual = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    Monthly.targets = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    Monthly.completion = defaultdict(lambda: defaultdict(dict))

    # 构造 person_data 按周→人，再按月份归属
    for dr in week_ranges:
        m = month_of_range(dr)
        if not m: continue
        m_key = str(m)
        for person, data in person_data.items():
            if dr in data:
                for metric in ["gsv", "profit", "real_sales", "qty"]:
                    Monthly.actual[person][m_key][metric] += data[dr][metric]
        if dr in gmv_person:
            for person, gmv in gmv_person[dr].items():
                Monthly.actual[person][m_key]["gmv"] += gmv

    # 读取 target
    for person, months in person_targets.items():
        for m_key, metrics in months.items():
            for key, val in metrics.items():
                Monthly.targets[person][m_key][key] = val

    # 计算
    today = datetime.now()
    today6 = datetime(today.year, today.month, today.day)

    for person in person_targets:
        Monthly.completion[person] = {}
        for m in range(2, 13):  # 2月到1月
            m_key = str(m)
            # 时间进度
            if m <= 1:
                year = 2027
            else:
                year = 2026
            month_start = datetime(year, m, 1)
            if m == 12:
                month_end = datetime(year, 12, 31)
            else:
                month_end = datetime(year, m + 1, 1) - timedelta(days=1)

            if today6 > month_end:
                time_progress = 100.0
            elif today6 < month_start:
                time_progress = 0.0
            else:
                time_progress = round((today6 - month_start).days / (month_end - month_start).days * 100, 1)

            a = Monthly.actual.get(person, {}).get(m_key, {})
            t = Monthly.targets.get(person, {}).get(m_key, {})

            entry = {
                "time_progress": time_progress,
                "actual": {
                    "profit": round(a.get("profit", 0), 2),
                    "gsv": round(a.get("gsv", 0), 2),
                    "real_sales": round(a.get("real_sales", 0), 2),
                    "qty": int(a.get("qty", 0)),
                    "gmv": round(a.get("gmv", 0), 2),
                },
                "target": {
                    "profit_target": round(t.get("profit_target", 0), 2),
                    "sales_target": int(t.get("sales_target", 0)),
                    "real_sales_target": round(t.get("real_sales_target", 0), 2),
                    "gmv_target": round(t.get("gmv_target", 0), 2),
                    "gsv_target": round(t.get("gsv_target", 0), 2),
                    "margin_target": t.get("margin_target", 15.0),
                    "return_target": t.get("return_target", 65.0),
                },
            }

            # 计算派生指标
            a_profit = a.get("profit", 0)
            a_gsv = a.get("gsv", 0)
            a_real = a.get("real_sales", 0)
            a_gmv = a.get("gmv", 0)

            entry["derived"] = {
                "margin_pct": round(a_profit / a_gsv * 100, 2) if a_gsv > 0 else "-",
                "return_rate": round((1 - (a_gsv * 12.5) / a_gmv) * 100, 2) if a_gmv > 0 else "-",
            }

            # 完成进度 = 完成率 / 时间进度
            def progress(actual, target_val):
                if target_val <= 0: return "-"
                if time_progress <= 0: return "-"
                rate = (actual / target_val) * 100
                return round(rate / time_progress * 100, 1)

            entry["completion"] = {
                "profit_progress": progress(a_profit, t.get("profit_target", 0)),
                "sales_progress": progress(a.get("qty", 0), t.get("sales_target", 0)),
                "real_sales_progress": progress(a_real, t.get("real_sales_target", 0)),
                "gmv_progress": progress(a_gmv, t.get("gmv_target", 0)),
                "gsv_progress": progress(a_gsv, t.get("gsv_target", 0)),
            }

            Monthly.completion[person][m_key] = entry

    return Monthly.completion

# ============================================================
# 主程序
# ============================================================

def main():
    print("="*60)
    print("build_data_v2.py - 男装版重构")
    print("="*60)

    print("\n[1/5] 从利润表A列生成周数组...")
    weeks, weeks_iso, week_labels, week_ranges, wy, wm, wym = generate_weeks_from_profit()

    print("\n[2/5] 读取产品列表...")
    sku_img, sku_owner, sku_first_date, sku_wb_id = read_product_list()

    print("\n[3/5] 读取年规目标...")
    person_targets = read_targets()

    print("\n[4/5] 读取利润表实际数据 + GMV...")
    shop_weekly, week_data, person_data = read_profit_data(week_ranges)
    traffic_weekly, gmv_person = read_traffic_and_gmv(week_ranges)

    print("\n[5/5] 月份归属 & 完成进度计算...")
    monthly_completion = compute_monthly(person_data, gmv_person, person_targets, week_ranges, wm)

    # 过滤掉2024年的周数据（只保留2025年及以后）
    keep_indices = [i for i, yr in enumerate(wy) if yr >= 2025]
    weeks2 = [weeks[i] for i in keep_indices]
    weeks_iso2 = [weeks_iso[i] for i in keep_indices]
    week_labels2 = [week_labels[i] for i in keep_indices]
    week_ranges2 = [week_ranges[i] for i in keep_indices]
    wy2 = [wy[i] for i in keep_indices]
    wm2 = [wm[i] for i in keep_indices]
    wym2 = [wym[i] for i in keep_indices]

    # 只保留2025+的数据
    shop_weekly2 = {}
    for shop, weeks_data in shop_weekly.items():
        fw = {k: v for k, v in weeks_data.items() if k in week_ranges2}
        if fw: shop_weekly2[shop] = fw

    week_data2 = {k: v for k, v in week_data.items() if k in week_ranges2}
    traffic_weekly2 = {k: v for k, v in traffic_weekly.items() if k in week_ranges2}

    data = {
        "WEEKS": weeks2,
        "WEEKS_ISO": weeks_iso2,
        "WEEK_LABELS": week_labels2,
        "WEEK_RANGES": week_ranges2,
        "WEEK_YEARS": wy2,
        "WEEK_MONTHS": wm2,
        "WEEK_YEAR_MONTHS": wym2,
        "SHOP_WEEKLY": shop_weekly2,
        "WEEK_DATA": week_data2,
        "TRAFFIC_WEEKLY": traffic_weekly2,
        "PERSON_TARGETS": person_targets,
        "MONTHLY_COMPLETION": monthly_completion,
        "SKU_IMG": sku_img,
        "SKU_OWNER": sku_owner,
        "SKU_FIRST_DATE": sku_first_date,
        "SKU_WB_ID": sku_wb_id,
        "_VERSION": "v2.0",
    }

    out_path = os.path.join(SRC_DIR, "data_v2_men.js")

    # Sanitize NaN/Infinity
    def sanitize(obj):
        if isinstance(obj, float):
            if math.isnan(obj) or math.isinf(obj): return None
            return obj
        if isinstance(obj, dict): return {k: sanitize(v) for k, v in obj.items()}
        if isinstance(obj, list): return [sanitize(v) for v in obj]
        return obj

    data_clean = sanitize(data)
    json_str = json.dumps(data_clean, ensure_ascii=False, default=str)
    js_content = f"var DATA_V2 = {json_str};"

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(js_content)

    size_mb = os.path.getsize(out_path) / 1024 / 1024
    print(f"\n输出: {out_path} ({size_mb:.1f} MB)")

    print(f"\n验证:")
    print(f"  周数: {len(weeks2)} (2025+，过滤前 {len(weeks)})")
    print(f"  目标人数: {len(person_targets)}")
    print(f"  店铺数: {len(shop_weekly2)}")
    print(f"  产品 SKU: {len(sku_img)}")
    print(f"\n完成!")

if __name__ == "__main__":
    main()
